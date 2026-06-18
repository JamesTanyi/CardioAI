#!/usr/bin/env python
"""
BloodTrack CloudRun 主入口
"""

import os
import sys
import json
from flask import Flask, request, jsonify
from datetime import datetime

# 数据库配置
# 优先级: 强制 SQLite > 环境变量 USE_CLOUD_DB > 默认 MySQL
_force_sqlite = os.environ.get("FORCE_SQLITE", "").lower() == "true"
_use_cloud_db_env = os.environ.get("USE_CLOUD_DB", "true").lower() == "true"
USE_CLOUD_DB = _use_cloud_db_env and not _force_sqlite

if USE_CLOUD_DB:
    import pymysql
    DB_CONFIG = {
        'host': os.environ.get("DB_HOST", "10.0.0.100"),
        'port': int(os.environ.get("DB_PORT", 3306)),
        'user': os.environ.get("DB_USER", "root"),
        'password': os.environ.get("DB_PASSWORD", ""),
        'database': os.environ.get("DB_NAME", "cardioai"),
        'charset': 'utf8mb4',
        'connect_timeout': 5
    }
    print("🔄 尝试连接腾讯云 MySQL 数据库...", flush=True)
    
    def _get_cloud_db():
        conn = pymysql.connect(**DB_CONFIG)
        conn.cursorclass = pymysql.cursors.DictCursor
        return conn
    
    def _init_cloud_db():
        conn = _get_cloud_db()
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS measurements (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                user_id     VARCHAR(100) NOT NULL,
                sbp         INT NOT NULL,
                dbp         INT NOT NULL,
                hr          INT DEFAULT 75,
                symptoms    TEXT DEFAULT '[]',
                risk_level  VARCHAR(20) DEFAULT 'normal',
                risk_text   TEXT DEFAULT '',
                analysis    TEXT DEFAULT '{}',
                datetime    VARCHAR(50) NOT NULL,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_user_id (user_id),
                INDEX idx_datetime (datetime)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                user_id     VARCHAR(100) UNIQUE NOT NULL,
                name        VARCHAR(50) DEFAULT '',
                age         INT DEFAULT 0,
                gender      VARCHAR(10) DEFAULT '',
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS family_bindings (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                family_id   VARCHAR(100) NOT NULL,
                patient_id  VARCHAR(100) NOT NULL,
                name        VARCHAR(50) NOT NULL,
                status      VARCHAR(20) DEFAULT 'active',
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY unique_binding (family_id, patient_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS feedbacks (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                from_id     VARCHAR(100) NOT NULL,
                from_role   VARCHAR(20) NOT NULL,
                to_id       VARCHAR(100) NOT NULL,
                content     TEXT NOT NULL,
                is_read     INT DEFAULT 0,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_to_id (to_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS doctor_bindings (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                doctor_id   VARCHAR(100) NOT NULL,
                patient_id  VARCHAR(100) NOT NULL,
                doctor_name VARCHAR(50) DEFAULT '',
                hospital    VARCHAR(200) DEFAULT '',
                department  VARCHAR(100) DEFAULT '',
                status      VARCHAR(20) DEFAULT 'active',
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY unique_dr_binding (doctor_id, patient_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS invite_codes (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                code        VARCHAR(10) NOT NULL UNIQUE,
                patient_id  VARCHAR(100) NOT NULL,
                used        INT DEFAULT 0,
                used_by     VARCHAR(100) DEFAULT '',
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at  TIMESTAMP NULL,
                INDEX idx_code (code),
                INDEX idx_patient (patient_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS invite_tokens (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                token       VARCHAR(64) NOT NULL UNIQUE,
                patient_id  VARCHAR(100) NOT NULL,
                role        VARCHAR(20) NOT NULL,
                used        INT DEFAULT 0,
                used_by     VARCHAR(100) DEFAULT '',
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at  TIMESTAMP NULL,
                INDEX idx_token (token),
                INDEX idx_patient (patient_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # ★ v8/v9 迁移：为存量表添加 status 字段 + invite_tokens 表
        for tbl in ['family_bindings', 'doctor_bindings']:
            try:
                cursor.execute(f"ALTER TABLE {tbl} ADD COLUMN status VARCHAR(20) DEFAULT 'active'")
            except Exception:
                pass  # 字段已存在
        conn.commit()
        conn.close()
        print("✅ [DB] MySQL 初始化完成", flush=True)
    
    # 尝试连接 MySQL，失败则自动降级到 SQLite
    try:
        _init_cloud_db()
        get_db = _get_cloud_db
        init_db = _init_cloud_db
    except Exception as mysql_err:
        print(f"❌ MySQL 连接失败: {mysql_err}", flush=True)
        print("🔽 自动降级到 SQLite 模式", flush=True)
        USE_CLOUD_DB = False

if not USE_CLOUD_DB:
    import sqlite3
    # ★ 使用持久化目录：优先环境变量 → /workspace/data/ → 降级 /tmp
    _default_db_dir = "/workspace/data"
    _env_db_path = os.environ.get("DB_PATH", "")
    if _env_db_path:
        DB_PATH = _env_db_path
    else:
        if os.path.isdir(_default_db_dir) and os.access(_default_db_dir, os.W_OK):
            DB_PATH = os.path.join(_default_db_dir, "bloodtrack.db")
        else:
            try:
                os.makedirs(_default_db_dir, exist_ok=True)
                DB_PATH = os.path.join(_default_db_dir, "bloodtrack.db")
            except Exception:
                DB_PATH = "/tmp/bloodtrack.db"
    # 确保 DB 目录存在
    _db_dir = os.path.dirname(DB_PATH)
    if _db_dir:
        os.makedirs(_db_dir, exist_ok=True)
    print(f"⚠️ 使用本地 SQLite 数据库: {DB_PATH}", flush=True)
    
    def get_db():
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        return conn
    
    def init_db():
        conn = get_db()
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS measurements (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id     TEXT NOT NULL,
                sbp         INTEGER NOT NULL,
                dbp         INTEGER NOT NULL,
                hr          INTEGER DEFAULT 75,
                symptoms    TEXT DEFAULT '[]',
                risk_level  TEXT DEFAULT 'normal',
                risk_text   TEXT DEFAULT '',
                analysis    TEXT DEFAULT '{}',
                datetime    TEXT NOT NULL,
                created_at  TEXT DEFAULT (datetime('now'))
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id     TEXT UNIQUE NOT NULL,
                name        TEXT DEFAULT '',
                age         INTEGER DEFAULT 0,
                gender      TEXT DEFAULT '',
                created_at  TEXT DEFAULT (datetime('now'))
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS family_bindings (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                family_id   TEXT NOT NULL,
                patient_id  TEXT NOT NULL,
                name        TEXT NOT NULL,
                status      TEXT DEFAULT 'active',
                created_at  TEXT DEFAULT (datetime('now')),
                UNIQUE(family_id, patient_id)
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS feedbacks (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                from_id     TEXT NOT NULL,
                from_role   TEXT NOT NULL,
                to_id       TEXT NOT NULL,
                content     TEXT NOT NULL,
                is_read     INTEGER DEFAULT 0,
                created_at  TEXT DEFAULT (datetime('now'))
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS doctor_bindings (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                doctor_id   TEXT NOT NULL,
                patient_id  TEXT NOT NULL,
                doctor_name TEXT DEFAULT '',
                hospital    TEXT DEFAULT '',
                department  TEXT DEFAULT '',
                status      TEXT DEFAULT 'active',
                created_at  TEXT DEFAULT (datetime('now')),
                UNIQUE(doctor_id, patient_id)
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS invite_codes (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                code        TEXT NOT NULL UNIQUE,
                patient_id  TEXT NOT NULL,
                used        INTEGER DEFAULT 0,
                used_by     TEXT DEFAULT '',
                created_at  TEXT DEFAULT (datetime('now')),
                expires_at  TEXT
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS invite_tokens (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                token       TEXT NOT NULL UNIQUE,
                patient_id  TEXT NOT NULL,
                role        TEXT NOT NULL,
                used        INTEGER DEFAULT 0,
                used_by     TEXT DEFAULT '',
                created_at  TEXT DEFAULT (datetime('now')),
                expires_at  TEXT
            )
        """)
        # ★ v8 迁移：为存量表添加 status 字段
        for tbl in ['family_bindings', 'doctor_bindings']:
            try:
                c.execute(f"ALTER TABLE {tbl} ADD COLUMN status TEXT DEFAULT 'active'")
            except Exception:
                pass  # 字段已存在
        conn.commit()
        conn.close()
        print("✅ [DB] SQLite 初始化完成", flush=True)

try:
    from engine.cardiovascular_engine import CardiovascularEngine
    EngineClass = CardiovascularEngine
    EngineError = None
except Exception as e:
    print(f"❌ 警告: 无法导入 CardiovascularEngine: {e}", flush=True)
    EngineClass = None
    EngineError = str(e)

def _normalize_record_time(rec):
    if rec is None:
        return rec
    rec = dict(rec)
    ts = rec.get("datetime") or rec.get("timestamp") or rec.get("date")
    if isinstance(ts, str):
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M"):
            try:
                rec["datetime"] = datetime.strptime(ts, fmt)
                break
            except ValueError:
                continue
    if "datetime" not in rec or not isinstance(rec["datetime"], datetime):
        rec["datetime"] = datetime.now()
    if "sbp" in rec and "dbp" in rec:
        rec["pp"] = rec.get("pp", rec["sbp"] - rec["dbp"])
    elif "pp" not in rec:
        rec["pp"] = 40
    if "hr" not in rec:
        rec["hr"] = 70
    return rec


def _format_record_for_db(item):
    return {
        "user_id": item.get("userId") or item.get("user_id"),
        "sbp": int(item.get("sbp", 0)),
        "dbp": int(item.get("dbp", 0)),
        "hr": int(item.get("hr", 75)),
        "symptoms": item.get("symptoms", []),
        "risk_level": item.get("riskLevel") or item.get("risk_level") or "normal",
        "risk_text": item.get("riskText") or item.get("risk_text") or "",
        "analysis": item.get("analysis") or {},
        "datetime": item.get("datetime") or item.get("date") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }


def _save_measurement(conn, cursor, item):
    record = _format_record_for_db(item)
    if not all([record["user_id"], record["sbp"], record["dbp"], record["datetime"]]):
        raise ValueError("Missing required record fields")
    if USE_CLOUD_DB:
        cursor.execute("INSERT IGNORE INTO users (user_id) VALUES (%s)", (record["user_id"],))
        cursor.execute(
            """
                INSERT INTO measurements
                    (user_id, sbp, dbp, hr, symptoms, risk_level, risk_text, analysis, datetime)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                record["user_id"],
                record["sbp"], record["dbp"], record["hr"],
                json.dumps(record["symptoms"], ensure_ascii=False),
                record["risk_level"], record["risk_text"],
                json.dumps(record["analysis"], ensure_ascii=False),
                record["datetime"]
            )
        )
    else:
        cursor.execute("INSERT OR IGNORE INTO users (user_id) VALUES (?)", (record["user_id"],))
        cursor.execute(
            """
                INSERT INTO measurements
                    (user_id, sbp, dbp, hr, symptoms, risk_level, risk_text, analysis, datetime)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                record["user_id"],
                record["sbp"], record["dbp"], record["hr"],
                json.dumps(record["symptoms"], ensure_ascii=False),
                record["risk_level"], record["risk_text"],
                json.dumps(record["analysis"], ensure_ascii=False),
                record["datetime"]
            )
        )


def _fetch_history_from_db(user_id, limit=90):
    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT * FROM measurements WHERE user_id=%s ORDER BY datetime DESC LIMIT %s",
                (user_id, limit)
            )
        else:
            cursor.execute(
                "SELECT * FROM measurements WHERE user_id = ? ORDER BY datetime DESC LIMIT ?",
                (user_id, limit)
            )
        rows = cursor.fetchall()
        results = []
        for row in rows:
            rec = dict(row) if isinstance(row, dict) else dict(row)
            rec["symptoms"] = json.loads(rec.get("symptoms") or "[]")
            rec["analysis"] = json.loads(rec.get("analysis") or "{}")
            results.append({
                "userId": rec.get("user_id"),
                "sbp": rec.get("sbp"),
                "dbp": rec.get("dbp"),
                "hr": rec.get("hr"),
                "symptoms": rec.get("symptoms"),
                "riskLevel": rec.get("risk_level"),
                "riskText": rec.get("risk_text"),
                "analysis": rec.get("analysis"),
                "datetime": rec.get("datetime")
            })
        return results
    finally:
        conn.close()

app = Flask(__name__)
init_db()

@app.route("/", methods=["GET"])
def health():
    return "Python service is running"

# ──────────────────────────────────────────────
# /analyze  核心分析
# ──────────────────────────────────────────────
@app.route("/analyze", methods=["POST"])
def analyze():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    print("📥 [Request] 收到 /analyze 请求", flush=True)
    if not data:
        return jsonify({"error": "Empty request body"}), 400
    if EngineClass is None:
        return jsonify({"code": -1, "error": "Engine load failed", "detail": EngineError}), 500

    history = data.get("history", [])
    current = data.get("current")
    if current is None:
        return jsonify({"error": "Missing 'current' record"}), 400
    if history is None:
        history = []
    if not isinstance(history, list):
        return jsonify({"error": "'history' must be a list"}), 400

    if not history:
        fallback_user_id = current.get("userId") or current.get("user_id")
        if fallback_user_id:
            history = _fetch_history_from_db(fallback_user_id, limit=90)

    history = [_normalize_record_time(r) for r in history]
    current = _normalize_record_time(current)

    try:
        engine = EngineClass(history, current)
        result = engine.run_all_diagnostics()
        print(f"✅ [Engine] 风险等级: {result.get('risk_level')}", flush=True)
        return jsonify({"code": 0, "data": result})
    except Exception as e:
        return jsonify({"error": "Engine execution failed", "detail": str(e)}), 500

# ──────────────────────────────────────────────
# /save_history  保存测量记录
# ──────────────────────────────────────────────
@app.route("/save_history", methods=["POST"])
def save_history():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    records = []
    if isinstance(data, dict) and data.get("history") and isinstance(data.get("history"), list):
        records = data.get("history")
    elif isinstance(data, list):
        records = data
    else:
        records = [data]

    conn = get_db()
    cursor = conn.cursor()
    saved = 0
    try:
        for item in records:
            try:
                _save_measurement(conn, cursor, item)
                saved += 1
            except Exception as inner_e:
                print(f"⚠️ 跳过无效记录: {inner_e}", flush=True)

        conn.commit()
        return jsonify({"code": 0, "message": "保存成功", "saved": saved})
    except Exception as e:
        return jsonify({"error": "保存失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /get_history  读取历史（患者自己或已绑定家属/医生）
# ──────────────────────────────────────────────
@app.route("/get_history", methods=["GET"])
def get_history():
    user_id   = request.args.get("userId") or request.args.get("user_id")
    viewer_id = request.args.get("viewerId")
    limit     = int(request.args.get("limit", 90))

    if not user_id:
        return jsonify({"error": "缺少 userId 参数"}), 400

    if viewer_id and viewer_id != user_id:
        conn = get_db()
        cursor = conn.cursor()
        # ★ v8：仅 active 绑定有权限查看数据
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT id FROM family_bindings WHERE family_id=%s AND patient_id=%s AND status='active'",
                (viewer_id, user_id)
            )
            family_binding = cursor.fetchone()
            if not family_binding:
                cursor.execute(
                    "SELECT id FROM doctor_bindings WHERE doctor_id=%s AND patient_id=%s AND status='active'",
                    (viewer_id, user_id)
                )
                doctor_binding = cursor.fetchone()
            else:
                doctor_binding = None
        else:
            cursor.execute(
                "SELECT id FROM family_bindings WHERE family_id=? AND patient_id=? AND status='active'",
                (viewer_id, user_id)
            )
            family_binding = cursor.fetchone()
            if not family_binding:
                cursor.execute(
                    "SELECT id FROM doctor_bindings WHERE doctor_id=? AND patient_id=? AND status='active'",
                    (viewer_id, user_id)
                )
                doctor_binding = cursor.fetchone()
            else:
                doctor_binding = None
        conn.close()
        if not family_binding and not doctor_binding:
            return jsonify({"error": "无权限查看该用户数据，请先确认绑定"}), 403

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute("""
                SELECT * FROM measurements
                WHERE user_id = %s
                ORDER BY datetime DESC
                LIMIT %s
            """, (user_id, limit))
        else:
            cursor.execute("""
                SELECT * FROM measurements
                WHERE user_id = ?
                ORDER BY datetime DESC
                LIMIT ?
            """, (user_id, limit))
        
        rows = cursor.fetchall()
        
        records = []
        for row in rows:
            rec = dict(row) if isinstance(row, dict) else dict(row)
            # MySQL 返回的是字符串，需要解析
            rec["symptoms"] = json.loads(rec.get("symptoms") or "[]")
            rec["analysis"] = json.loads(rec.get("analysis") or "{}")
            records.append(rec)

        return jsonify({"code": 0, "data": records})
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /bind_family  家属绑定患者
# ──────────────────────────────────────────────
@app.route("/bind_family", methods=["POST"])
def bind_family():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    family_id  = data.get("familyId")
    patient_id = data.get("patientId")
    name       = data.get("name", "家人")

    if not all([family_id, patient_id]):
        return jsonify({"error": "缺少 familyId 或 patientId"}), 400
    if family_id == patient_id:
        return jsonify({"error": "不能绑定自己"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute("SELECT id FROM users WHERE user_id=%s", (patient_id,))
        else:
            cursor.execute("SELECT id FROM users WHERE user_id=?", (patient_id,))
        user = cursor.fetchone()
        if not user:
            # 患者尚未向服务器同步过数据，自动注册
            if USE_CLOUD_DB:
                cursor.execute("INSERT IGNORE INTO users (user_id) VALUES (%s)", (patient_id,))
            else:
                cursor.execute("INSERT OR IGNORE INTO users (user_id) VALUES (?)", (patient_id,))
            print(f"📝 [DB] 自动创建患者账户: {patient_id}", flush=True)

        # ★ v8：检测当前绑定状态，如果是 active 则不覆盖
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT status FROM family_bindings WHERE family_id=%s AND patient_id=%s",
                (family_id, patient_id)
            )
        else:
            cursor.execute(
                "SELECT status FROM family_bindings WHERE family_id=? AND patient_id=?",
                (family_id, patient_id)
            )
        existing = cursor.fetchone()
        existing_status = existing["status"] if existing else None

        # 已 active → 幂等直接返回成功
        if existing_status == 'active':
            conn.close()
            return jsonify({"code": 0, "message": "绑定已存在", "status": "active"})

        # 新绑定或 pending 绑定 → 写为 pending（等待用户确认）
        if USE_CLOUD_DB:
            if existing:
                cursor.execute(
                    "UPDATE family_bindings SET name=%s, status='pending' WHERE family_id=%s AND patient_id=%s",
                    (name, family_id, patient_id)
                )
            else:
                cursor.execute("""
                    INSERT INTO family_bindings (family_id, patient_id, name, status)
                    VALUES (%s, %s, %s, 'pending')
                """, (family_id, patient_id, name))
        else:
            if existing:
                cursor.execute(
                    "UPDATE family_bindings SET name=?, status='pending' WHERE family_id=? AND patient_id=?",
                    (name, family_id, patient_id)
                )
            else:
                cursor.execute("""
                    INSERT INTO family_bindings (family_id, patient_id, name, status)
                    VALUES (?, ?, ?, 'pending')
                """, (family_id, patient_id, name))
        conn.commit()
        print(f"🔗 [DB] 家属绑定(pending): {family_id} → {patient_id} ({name})", flush=True)
        return jsonify({"code": 0, "message": "绑定已提交，等待确认", "status": "pending"})
    except Exception as e:
        return jsonify({"error": "绑定失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /get_family_list  获取已绑定家人列表
# ──────────────────────────────────────────────
@app.route("/get_family_list", methods=["GET"])
def get_family_list():
    family_id = request.args.get("familyId")
    if not family_id:
        return jsonify({"error": "缺少 familyId 参数"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute("""
                SELECT patient_id, name, created_at
                FROM family_bindings
                WHERE family_id = %s AND status='active'
                ORDER BY created_at DESC
            """, (family_id,))
        else:
            cursor.execute("""
                SELECT patient_id, name, created_at
                FROM family_bindings
                WHERE family_id = ? AND status='active'
                ORDER BY created_at DESC
            """, (family_id,))
        rows = cursor.fetchall()
        return jsonify({"code": 0, "data": [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /send_feedback  家属或医生发反馈给患者
# ──────────────────────────────────────────────
@app.route("/send_feedback", methods=["POST"])
def send_feedback():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    from_id   = data.get("fromId")
    from_role = data.get("fromRole", "family")
    to_id     = data.get("toId")
    content   = data.get("content", "").strip()

    if not all([from_id, to_id, content]):
        return jsonify({"error": "缺少 fromId / toId / content"}), 400
    if len(content) > 500:
        return jsonify({"error": "反馈内容不能超过500字"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        # 根据角色检查对应的绑定关系（仅 active）
        if from_role == "doctor":
            if USE_CLOUD_DB:
                cursor.execute(
                    "SELECT id FROM doctor_bindings WHERE doctor_id=%s AND patient_id=%s AND status='active'",
                    (from_id, to_id)
                )
            else:
                cursor.execute(
                    "SELECT id FROM doctor_bindings WHERE doctor_id=? AND patient_id=? AND status='active'",
                    (from_id, to_id)
                )
        else:
            if USE_CLOUD_DB:
                cursor.execute(
                    "SELECT id FROM family_bindings WHERE family_id=%s AND patient_id=%s AND status='active'",
                    (from_id, to_id)
                )
            else:
                cursor.execute(
                    "SELECT id FROM family_bindings WHERE family_id=? AND patient_id=? AND status='active'",
                    (from_id, to_id)
                )
        binding = cursor.fetchone()
        if not binding:
            return jsonify({"error": "未绑定该患者，无法发送反馈"}), 403

        if USE_CLOUD_DB:
            cursor.execute("""
                INSERT INTO feedbacks (from_id, from_role, to_id, content)
                VALUES (%s, %s, %s, %s)
            """, (from_id, from_role, to_id, content))
        else:
            cursor.execute("""
                INSERT INTO feedbacks (from_id, from_role, to_id, content)
                VALUES (?, ?, ?, ?)
            """, (from_id, from_role, to_id, content))
        conn.commit()
        print(f"💬 [DB] 反馈: {from_id}({from_role}) → {to_id}", flush=True)
        return jsonify({"code": 0, "message": "反馈已发送"})
    except Exception as e:
        return jsonify({"error": "发送失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /get_feedback  患者查看收到的反馈
# ──────────────────────────────────────────────
@app.route("/get_feedback", methods=["GET"])
def get_feedback():
    user_id = request.args.get("userId")
    if not user_id:
        return jsonify({"error": "缺少 userId 参数"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute("""
                SELECT id, from_id, from_role, content, is_read, created_at
                FROM feedbacks
                WHERE to_id = %s
                ORDER BY created_at DESC
                LIMIT 50
            """, (user_id,))
        else:
            cursor.execute("""
                SELECT id, from_id, from_role, content, is_read, created_at
                FROM feedbacks
                WHERE to_id = ?
                ORDER BY created_at DESC
                LIMIT 50
            """, (user_id,))
        rows = cursor.fetchall()

        feedbacks = [dict(r) for r in rows]
        unread_count = sum(1 for f in feedbacks if f["is_read"] == 0)

        if USE_CLOUD_DB:
            cursor.execute(
                "UPDATE feedbacks SET is_read=1 WHERE to_id=%s AND is_read=0",
                (user_id,)
            )
        else:
            cursor.execute(
                "UPDATE feedbacks SET is_read=1 WHERE to_id=? AND is_read=0",
                (user_id,)
            )
        conn.commit()
        return jsonify({"code": 0, "data": feedbacks, "unread": unread_count})
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /bind_doctor  医生绑定患者
# ──────────────────────────────────────────────
@app.route("/bind_doctor", methods=["POST"])
def bind_doctor():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    doctor_id   = data.get("doctorId")
    patient_id  = data.get("patientId")
    doctor_name = data.get("doctorName", "")
    hospital    = data.get("hospital", "")
    department  = data.get("department", "")

    if not all([doctor_id, patient_id]):
        return jsonify({"error": "缺少 doctorId / patientId"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT id FROM users WHERE user_id=%s", (patient_id,)
            )
        else:
            cursor.execute(
                "SELECT id FROM users WHERE user_id=?", (patient_id,)
            )
        user = cursor.fetchone()
        if not user:
            # 患者尚未向服务器同步过数据，自动注册
            if USE_CLOUD_DB:
                cursor.execute("INSERT IGNORE INTO users (user_id) VALUES (%s)", (patient_id,))
            else:
                cursor.execute("INSERT OR IGNORE INTO users (user_id) VALUES (?)", (patient_id,))
            print(f"📝 [DB] 自动创建患者账户: {patient_id}", flush=True)

        # ★ v8：检测当前绑定状态，活跃绑定不覆盖
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT status FROM doctor_bindings WHERE doctor_id=%s AND patient_id=%s",
                (doctor_id, patient_id)
            )
        else:
            cursor.execute(
                "SELECT status FROM doctor_bindings WHERE doctor_id=? AND patient_id=?",
                (doctor_id, patient_id)
            )
        existing = cursor.fetchone()
        existing_status = existing["status"] if existing else None

        if existing_status == 'active':
            conn.close()
            return jsonify({"code": 0, "message": "绑定已存在", "status": "active"})

        if USE_CLOUD_DB:
            if existing:
                cursor.execute(
                    "UPDATE doctor_bindings SET doctor_name=%s, hospital=%s, department=%s, status='pending' WHERE doctor_id=%s AND patient_id=%s",
                    (doctor_name, hospital, department, doctor_id, patient_id)
                )
            else:
                cursor.execute("""
                    INSERT INTO doctor_bindings (doctor_id, patient_id, doctor_name, hospital, department, status)
                    VALUES (%s, %s, %s, %s, %s, 'pending')
                """, (doctor_id, patient_id, doctor_name, hospital, department))
        else:
            if existing:
                cursor.execute(
                    "UPDATE doctor_bindings SET doctor_name=?, hospital=?, department=?, status='pending' WHERE doctor_id=? AND patient_id=?",
                    (doctor_name, hospital, department, doctor_id, patient_id)
                )
            else:
                cursor.execute("""
                    INSERT INTO doctor_bindings (doctor_id, patient_id, doctor_name, hospital, department, status)
                    VALUES (?, ?, ?, ?, ?, 'pending')
                """, (doctor_id, patient_id, doctor_name, hospital, department))
        conn.commit()
        print(f"🩺 [DB] 医生绑定(pending): {doctor_id} → {patient_id} ({doctor_name})", flush=True)
        return jsonify({"code": 0, "message": "绑定已提交，等待确认", "status": "pending"})
    except Exception as e:
        return jsonify({"error": "绑定失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# ★ v8 /confirm_binding  家属/医生确认绑定（pending → active）
# ──────────────────────────────────────────────
@app.route("/confirm_binding", methods=["POST"])
def confirm_binding():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    viewer_id = data.get("viewerId")  # 家属ID 或 医生ID
    patient_id = data.get("patientId")
    role = data.get("role", "family")  # 'family' 或 'doctor'

    if not all([viewer_id, patient_id]):
        return jsonify({"error": "缺少 viewerId / patientId"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if role == 'doctor':
            table = 'doctor_bindings'
            viewer_col = 'doctor_id'
        else:
            table = 'family_bindings'
            viewer_col = 'family_id'

        if USE_CLOUD_DB:
            cursor.execute(
                f"SELECT id, status FROM {table} WHERE {viewer_col}=%s AND patient_id=%s",
                (viewer_id, patient_id)
            )
        else:
            cursor.execute(
                f"SELECT id, status FROM {table} WHERE {viewer_col}=? AND patient_id=?",
                (viewer_id, patient_id)
            )
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "未找到待确认的绑定关系"}), 404

        current_status = row["status"] if isinstance(row, dict) else row["status"] if hasattr(row, "status") else "pending"
        if current_status == 'active':
            return jsonify({"code": 0, "message": "绑定已确认", "status": "active"})

        if USE_CLOUD_DB:
            cursor.execute(
                f"UPDATE {table} SET status='active' WHERE {viewer_col}=%s AND patient_id=%s",
                (viewer_id, patient_id)
            )
        else:
            cursor.execute(
                f"UPDATE {table} SET status='active' WHERE {viewer_col}=? AND patient_id=?",
                (viewer_id, patient_id)
            )
        conn.commit()
        print(f"✅ [DB] 绑定已确认: {viewer_id}({role}) → {patient_id}", flush=True)
        return jsonify({"code": 0, "message": "绑定已确认", "status": "active"})
    except Exception as e:
        return jsonify({"error": "确认失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# ★ v8 /reject_binding  家属/医生拒绝绑定（pending → rejected）
# ──────────────────────────────────────────────
@app.route("/reject_binding", methods=["POST"])
def reject_binding():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    viewer_id = data.get("viewerId")
    patient_id = data.get("patientId")
    role = data.get("role", "family")

    if not all([viewer_id, patient_id]):
        return jsonify({"error": "缺少 viewerId / patientId"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if role == 'doctor':
            table = 'doctor_bindings'
            viewer_col = 'doctor_id'
        else:
            table = 'family_bindings'
            viewer_col = 'family_id'

        if USE_CLOUD_DB:
            cursor.execute(
                f"UPDATE {table} SET status='rejected' WHERE {viewer_col}=%s AND patient_id=%s AND status='pending'",
                (viewer_id, patient_id)
            )
        else:
            cursor.execute(
                f"UPDATE {table} SET status='rejected' WHERE {viewer_col}=? AND patient_id=? AND status='pending'",
                (viewer_id, patient_id)
            )
        conn.commit()
        print(f"❌ [DB] 绑定已拒绝: {viewer_id}({role}) → {patient_id}", flush=True)
        return jsonify({"code": 0, "message": "已拒绝绑定"})
    except Exception as e:
        return jsonify({"error": "操作失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# ★ v8 /get_patient_summary  获取患者摘要信息（确认绑定时展示）
# ──────────────────────────────────────────────
@app.route("/get_patient_summary", methods=["GET"])
def get_patient_summary():
    patient_id = request.args.get("patientId", "").strip()
    if not patient_id:
        return jsonify({"error": "缺少 patientId"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        # 获取用户基本信息
        if USE_CLOUD_DB:
            cursor.execute("SELECT name, age, gender FROM users WHERE user_id=%s", (patient_id,))
        else:
            cursor.execute("SELECT name, age, gender FROM users WHERE user_id=?", (patient_id,))
        user = cursor.fetchone()

        # 获取最近一次测量
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT sbp, dbp, hr, risk_level, datetime FROM measurements WHERE user_id=%s ORDER BY datetime DESC LIMIT 1",
                (patient_id,)
            )
        else:
            cursor.execute(
                "SELECT sbp, dbp, hr, risk_level, datetime FROM measurements WHERE user_id=? ORDER BY datetime DESC LIMIT 1",
                (patient_id,)
            )
        latest = cursor.fetchone()

        # 获取总测量次数
        if USE_CLOUD_DB:
            cursor.execute("SELECT COUNT(*) as cnt FROM measurements WHERE user_id=%s", (patient_id,))
        else:
            cursor.execute("SELECT COUNT(*) as cnt FROM measurements WHERE user_id=?", (patient_id,))
        count_row = cursor.fetchone()

        result = {
            "patientId": patient_id,
            "name": user["name"] if user else "",
            "age": user["age"] if user else 0,
            "gender": user["gender"] if user else "",
            "totalMeasurements": count_row["cnt"] if count_row else 0,
            "latest": None
        }

        if latest:
            result["latest"] = {
                "sbp": latest["sbp"],
                "dbp": latest["dbp"],
                "hr": latest["hr"],
                "risk_level": latest["risk_level"],
                "datetime": str(latest["datetime"])[:16] if latest["datetime"] else ""
            }

        return jsonify({"code": 0, "data": result})
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# ★ v9 /generate_invite_token  生成通用邀请 Token（替代直接暴露 patientId）
# ──────────────────────────────────────────────
@app.route("/generate_invite_token", methods=["POST"])
def generate_invite_token():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    patient_id = data.get("patientId") or data.get("user_id")
    role = data.get("role", "family")  # 'family' 或 'doctor'

    if not patient_id:
        return jsonify({"error": "缺少 patientId"}), 400
    if role not in ('family', 'doctor'):
        return jsonify({"error": "role 必须是 family 或 doctor"}), 400

    import secrets

    conn = get_db()
    cursor = conn.cursor()
    try:
        token = secrets.token_hex(24)  # 48字符 hex 字符串

        if USE_CLOUD_DB:
            cursor.execute("""
                INSERT INTO invite_tokens (token, patient_id, role, expires_at)
                VALUES (%s, %s, %s, DATE_ADD(NOW(), INTERVAL 24 HOUR))
            """, (token, patient_id, role))
        else:
            cursor.execute("""
                INSERT INTO invite_tokens (token, patient_id, role, expires_at)
                VALUES (?, ?, ?, datetime('now', '+24 hours'))
            """, (token, patient_id, role))
        conn.commit()
        print(f"🔑 [DB] 生成邀请Token: {token[:8]}... → {patient_id} ({role})", flush=True)
        return jsonify({"code": 0, "data": {"token": token, "role": role, "expiresIn": "24小时"}})
    except Exception as e:
        return jsonify({"error": "生成失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# ★ v9 /validate_invite_token  验证邀请 Token（接收方打开链接时调用）
# ──────────────────────────────────────────────
@app.route("/validate_invite_token", methods=["POST"])
def validate_invite_token():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    token = data.get("token", "").strip()
    if not token:
        return jsonify({"error": "缺少 token"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        # 查找有效 token
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT * FROM invite_tokens WHERE token=%s",
                (token,)
            )
        else:
            cursor.execute(
                "SELECT * FROM invite_tokens WHERE token=?",
                (token,)
            )
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "邀请链接无效"}), 404

        row = dict(row)

        # 检查是否已使用
        if row.get("used"):
            return jsonify({"error": "该邀请链接已被使用"}), 400

        # 检查是否过期
        expires_str = row.get("expires_at")
        if expires_str:
            expires_at = datetime.strptime(str(expires_str)[:19], "%Y-%m-%d %H:%M:%S")
            if datetime.now() > expires_at:
                return jsonify({"error": "邀请链接已过期，请联系患者重新发送"}), 400

        patient_id = row["patient_id"]
        role = row.get("role", "family")

        # 获取患者摘要信息
        if USE_CLOUD_DB:
            cursor.execute("SELECT name, age, gender FROM users WHERE user_id=%s", (patient_id,))
        else:
            cursor.execute("SELECT name, age, gender FROM users WHERE user_id=?", (patient_id,))
        user = cursor.fetchone()

        # 获取最近一次测量
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT sbp, dbp, hr, risk_level, datetime FROM measurements WHERE user_id=%s ORDER BY datetime DESC LIMIT 1",
                (patient_id,)
            )
        else:
            cursor.execute(
                "SELECT sbp, dbp, hr, risk_level, datetime FROM measurements WHERE user_id=? ORDER BY datetime DESC LIMIT 1",
                (patient_id,)
            )
        latest = cursor.fetchone()

        # 获取总测量次数
        if USE_CLOUD_DB:
            cursor.execute("SELECT COUNT(*) as cnt FROM measurements WHERE user_id=%s", (patient_id,))
        else:
            cursor.execute("SELECT COUNT(*) as cnt FROM measurements WHERE user_id=?", (patient_id,))
        count_row = cursor.fetchone()

        patient_name = user["name"] if user else ""
        patient_info = {
            "name": patient_name,
            "age": user["age"] if user else 0,
            "gender": user["gender"] if user else "",
            "totalMeasurements": count_row["cnt"] if count_row else 0,
            "latest": None
        }

        if latest:
            patient_info["latest"] = {
                "sbp": latest["sbp"],
                "dbp": latest["dbp"],
                "hr": latest["hr"],
                "risk_level": latest["risk_level"],
                "datetime": str(latest["datetime"])[:16] if latest["datetime"] else ""
            }

        print(f"✅ [DB] Token验证成功: {token[:8]}... → {patient_id} ({role})", flush=True)
        return jsonify({
            "code": 0,
            "data": {
                "patientId": patient_id,
                "patientName": patient_name,
                "role": role,
                "patientSummary": patient_info
            }
        })
    except Exception as e:
        return jsonify({"error": "验证失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# ★ v9 /bind_by_token  通过邀请 Token 一步完成绑定（写入 status='active'）
#    接收方在确认页点击"确认绑定"时调用。验证 token → 写入 active 绑定 → 标记 token 已用。
# ──────────────────────────────────────────────
@app.route("/bind_by_token", methods=["POST"])
def bind_by_token():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    token = data.get("token", "").strip()
    viewer_id = data.get("viewerId", "").strip()
    viewer_name = data.get("viewerName", "").strip()
    hospital = data.get("hospital", "").strip()
    department = data.get("department", "").strip()

    if not token:
        return jsonify({"error": "缺少 token"}), 400
    if not viewer_id:
        return jsonify({"error": "缺少 viewerId"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        # 1. 查找 token
        if USE_CLOUD_DB:
            cursor.execute("SELECT * FROM invite_tokens WHERE token=%s", (token,))
        else:
            cursor.execute("SELECT * FROM invite_tokens WHERE token=?", (token,))
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "邀请链接无效"}), 404

        row = dict(row)

        # 2. 检查是否已使用
        if row.get("used"):
            return jsonify({"error": "该邀请链接已被使用"}), 400

        # 3. 检查是否过期
        expires_str = row.get("expires_at")
        if expires_str:
            expires_at = datetime.strptime(str(expires_str)[:19], "%Y-%m-%d %H:%M:%S")
            if datetime.now() > expires_at:
                return jsonify({"error": "邀请链接已过期，请联系患者重新发送"}), 400

        patient_id = row["patient_id"]
        role = row.get("role", "family")

        # 4. 自绑定守卫
        if viewer_id == patient_id:
            return jsonify({"error": "不能绑定自己"}), 400

        # 5. 根据 role 选表
        if role == "doctor":
            table = "doctor_bindings"
            viewer_col = "doctor_id"
            viewer_name_col = "doctor_name"
        else:
            table = "family_bindings"
            viewer_col = "family_id"
            viewer_name_col = "name"

        # 6. 查询现有绑定
        if USE_CLOUD_DB:
            cursor.execute(
                f"SELECT status FROM {table} WHERE {viewer_col}=%s AND patient_id=%s",
                (viewer_id, patient_id)
            )
        else:
            cursor.execute(
                f"SELECT status FROM {table} WHERE {viewer_col}=? AND patient_id=?",
                (viewer_id, patient_id)
            )
        existing = cursor.fetchone()

        # 7. 若已 active → 幂等返回
        #    MySQL DictCursor 返回 dict，SQLite Row 同样支持 ["status"] 访问
        existing_status = existing["status"] if existing else None
        if existing_status == "active":
            # 仍标记 token 已用（防止重复使用）
            if USE_CLOUD_DB:
                cursor.execute(
                    "UPDATE invite_tokens SET used=1, used_by=%s WHERE token=%s",
                    (viewer_id, token)
                )
            else:
                cursor.execute(
                    "UPDATE invite_tokens SET used=1, used_by=? WHERE token=?",
                    (viewer_id, token)
                )
            conn.commit()
            return jsonify({
                "code": 0,
                "status": "active",
                "message": "已绑定",
                "patientId": patient_id,
                "role": role
            })

        # 8. 一步写入 active（INSERT 或 UPDATE，不走 pending）
        if existing:
            if role == "doctor":
                if USE_CLOUD_DB:
                    cursor.execute(
                        f"UPDATE {table} SET {viewer_name_col}=%s, hospital=%s, department=%s, status='active' WHERE {viewer_col}=%s AND patient_id=%s",
                        (viewer_name or "医生", hospital, department, viewer_id, patient_id)
                    )
                else:
                    cursor.execute(
                        f"UPDATE {table} SET {viewer_name_col}=?, hospital=?, department=?, status='active' WHERE {viewer_col}=? AND patient_id=?",
                        (viewer_name or "医生", hospital, department, viewer_id, patient_id)
                    )
            else:
                if USE_CLOUD_DB:
                    cursor.execute(
                        f"UPDATE {table} SET {viewer_name_col}=%s, status='active' WHERE {viewer_col}=%s AND patient_id=%s",
                        (viewer_name or "家人", viewer_id, patient_id)
                    )
                else:
                    cursor.execute(
                        f"UPDATE {table} SET {viewer_name_col}=?, status='active' WHERE {viewer_col}=? AND patient_id=?",
                        (viewer_name or "家人", viewer_id, patient_id)
                    )
        else:
            if role == "doctor":
                if USE_CLOUD_DB:
                    cursor.execute(
                        f"INSERT INTO {table} (doctor_id, patient_id, doctor_name, hospital, department, status) VALUES (%s, %s, %s, %s, %s, 'active')",
                        (viewer_id, patient_id, viewer_name or "医生", hospital, department)
                    )
                else:
                    cursor.execute(
                        f"INSERT INTO {table} (doctor_id, patient_id, doctor_name, hospital, department, status) VALUES (?, ?, ?, ?, ?, 'active')",
                        (viewer_id, patient_id, viewer_name or "医生", hospital, department)
                    )
            else:
                if USE_CLOUD_DB:
                    cursor.execute(
                        f"INSERT INTO {table} (family_id, patient_id, name, status) VALUES (%s, %s, %s, 'active')",
                        (viewer_id, patient_id, viewer_name or "家人")
                    )
                else:
                    cursor.execute(
                        f"INSERT INTO {table} (family_id, patient_id, name, status) VALUES (?, ?, ?, 'active')",
                        (viewer_id, patient_id, viewer_name or "家人")
                    )

        # 9. 标记 token 已用
        if USE_CLOUD_DB:
            cursor.execute(
                "UPDATE invite_tokens SET used=1, used_by=%s WHERE token=%s",
                (viewer_id, token)
            )
        else:
            cursor.execute(
                "UPDATE invite_tokens SET used=1, used_by=? WHERE token=?",
                (viewer_id, token)
            )

        conn.commit()
        print(f"✅ [DB] Token绑定成功: {token[:8]}... → {viewer_id} 绑定患者 {patient_id} ({role})", flush=True)
        return jsonify({
            "code": 0,
            "status": "active",
            "message": "绑定成功",
            "patientId": patient_id,
            "role": role
        })
    except Exception as e:
        return jsonify({"error": "绑定失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /generate_invite_code  患者生成邀请码（用于医生/家人绑定）
# ──────────────────────────────────────────────
@app.route("/generate_invite_code", methods=["POST"])
def generate_invite_code():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    patient_id = data.get("patientId") or data.get("user_id")
    if not patient_id:
        return jsonify({"error": "缺少 patientId"}), 400

    import random
    import string

    conn = get_db()
    cursor = conn.cursor()
    try:
        # 生成 6 位随机数字
        code = ''.join(random.choices(string.digits, k=6))

        if USE_CLOUD_DB:
            cursor.execute("""
                INSERT INTO invite_codes (code, patient_id, expires_at)
                VALUES (%s, %s, DATE_ADD(NOW(), INTERVAL 24 HOUR))
            """, (code, patient_id))
        else:
            cursor.execute("""
                INSERT INTO invite_codes (code, patient_id, expires_at)
                VALUES (?, ?, datetime('now', '+24 hours'))
            """, (code, patient_id))
        conn.commit()
        print(f"📨 [DB] 生成邀请码: {code} → {patient_id}", flush=True)
        return jsonify({"code": 0, "data": {"inviteCode": code, "expiresIn": "24小时"}})
    except Exception as e:
        return jsonify({"error": "生成失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /bind_doctor_by_code  医生通过邀请码绑定患者
# ──────────────────────────────────────────────
@app.route("/bind_doctor_by_code", methods=["POST"])
def bind_doctor_by_code():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    invite_code = data.get("code", "").strip()
    doctor_id   = data.get("doctorId")
    doctor_name = data.get("doctorName", "")
    hospital    = data.get("hospital", "")
    department  = data.get("department", "")

    if not invite_code or not doctor_id:
        return jsonify({"error": "缺少 code / doctorId"}), 400
    if len(invite_code) != 6 or not invite_code.isdigit():
        return jsonify({"error": "邀请码格式不正确，应为6位数字"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT * FROM invite_codes WHERE code=%s",
                (invite_code,)
            )
        else:
            cursor.execute(
                "SELECT * FROM invite_codes WHERE code=?",
                (invite_code,)
            )
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "邀请码不存在"}), 404

        row = dict(row)
        if row.get("used"):
            return jsonify({"error": "邀请码已被使用"}), 400

        # 检查是否过期
        expires_str = row.get("expires_at")
        if expires_str:
            from datetime import datetime as dt
            expires_at = dt.strptime(str(expires_str), "%Y-%m-%d %H:%M:%S")
            if dt.now() > expires_at:
                return jsonify({"error": "邀请码已过期，请重新生成"}), 400

        patient_id = row["patient_id"]

        # 防止自己绑定自己
        if doctor_id == patient_id:
            return jsonify({"error": "不能绑定自己"}), 400

        # ★ v8：检查是否已 active 绑定，否则写为 pending
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT status FROM doctor_bindings WHERE doctor_id=%s AND patient_id=%s",
                (doctor_id, patient_id)
            )
        else:
            cursor.execute(
                "SELECT status FROM doctor_bindings WHERE doctor_id=? AND patient_id=?",
                (doctor_id, patient_id)
            )
        existing = cursor.fetchone()
        if existing and (existing["status"] if isinstance(existing, dict) else existing[0]) == 'active':
            conn.close()
            return jsonify({"code": 0, "message": "绑定已存在", "status": "active"})

        # 写入医生绑定表（pending）
        if USE_CLOUD_DB:
            if existing:
                cursor.execute(
                    "UPDATE doctor_bindings SET doctor_name=%s, hospital=%s, department=%s, status='pending' WHERE doctor_id=%s AND patient_id=%s",
                    (doctor_name, hospital, department, doctor_id, patient_id)
                )
            else:
                cursor.execute("""
                    INSERT INTO doctor_bindings (doctor_id, patient_id, doctor_name, hospital, department, status)
                    VALUES (%s, %s, %s, %s, %s, 'pending')
                """, (doctor_id, patient_id, doctor_name, hospital, department))
            # 标记邀请码已使用
            cursor.execute(
                "UPDATE invite_codes SET used=1, used_by=%s WHERE code=%s",
                (doctor_id, invite_code)
            )
        else:
            if existing:
                cursor.execute(
                    "UPDATE doctor_bindings SET doctor_name=?, hospital=?, department=?, status='pending' WHERE doctor_id=? AND patient_id=?",
                    (doctor_name, hospital, department, doctor_id, patient_id)
                )
            else:
                cursor.execute("""
                    INSERT INTO doctor_bindings (doctor_id, patient_id, doctor_name, hospital, department, status)
                    VALUES (?, ?, ?, ?, ?, 'pending')
                """, (doctor_id, patient_id, doctor_name, hospital, department))
            cursor.execute(
                "UPDATE invite_codes SET used=1, used_by=? WHERE code=?",
                (doctor_id, invite_code)
            )
        conn.commit()
        print(f"🩺 [DB] 医生通过邀请码绑定(pending): {doctor_id} → {patient_id} (code: {invite_code})", flush=True)
        return jsonify({"code": 0, "message": "绑定已提交，等待确认", "status": "pending", "patientId": patient_id, "patientName": row.get("patient_id", "")})
    except Exception as e:
        return jsonify({"error": "绑定失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /get_doctor_patients  医生查看已绑定患者
# ──────────────────────────────────────────────
@app.route("/get_doctor_patients", methods=["GET"])
def get_doctor_patients():
    doctor_id = request.args.get("doctorId")
    if not doctor_id:
        return jsonify({"error": "缺少 doctorId 参数"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute("""
                SELECT patient_id, doctor_name, hospital, department, created_at
                FROM doctor_bindings
                WHERE doctor_id = %s AND status='active'
                ORDER BY created_at DESC
            """, (doctor_id,))
        else:
            cursor.execute("""
                SELECT patient_id, doctor_name, hospital, department, created_at
                FROM doctor_bindings
                WHERE doctor_id = ? AND status='active'
                ORDER BY created_at DESC
            """, (doctor_id,))
        rows = cursor.fetchall()
        return jsonify({"code": 0, "data": [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /upload_excel  上传并解析 Excel 文件
# ──────────────────────────────────────────────
@app.route("/upload_excel", methods=["POST"])
def upload_excel():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "没有文件"}), 400
        
        file = request.files['file']
        user_id = request.form.get('userId') or request.form.get('user_id')
        
        if not user_id:
            return jsonify({"error": "缺少 userId"}), 400
        
        if file.filename == '':
            return jsonify({"error": "文件名为空"}), 400
        
        print(f"📥 收到 Excel 文件: {file.filename}", flush=True)
        
        # 读取文件内容
        file_content = file.read()
        
        # 使用 openpyxl 库解析
        try:
            import io
            from openpyxl import load_workbook
            
            wb = load_workbook(io.BytesIO(file_content))
            sheet = wb.active
            
            # 获取表头（第一行）
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).strip() if cell.value else '')
            
            print(f"📋 Excel 表头: {headers}", flush=True)
            
            # 解析数据行
            records = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(v is None for v in row):
                    continue
                
                record = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        record[headers[col_idx]] = value
                
                records.append(record)
            
            print(f"📊 解析到 {len(records)} 条记录", flush=True)
            
        except ImportError:
            return jsonify({"error": "服务器未安装 xlsx 库"}), 500
        except Exception as e:
            return jsonify({"error": f"Excel 解析失败: {str(e)}"}), 500
        
        if not records:
            return jsonify({"code": 0, "message": "文件中没有数据", "saved": 0})
        
        # 保存到数据库
        conn = get_db()
        cursor = conn.cursor()
        saved = 0
        
        try:
            for item in records:
                try:
                    record = _format_record_for_db({**item, "userId": user_id})
                    if not all([record["user_id"], record["sbp"], record["dbp"], record["datetime"]]):
                        continue
                    
                    _save_measurement(conn, cursor, {**item, "userId": user_id})
                    saved += 1
                except Exception as inner_e:
                    print(f"⚠️ 跳过无效记录: {inner_e}", flush=True)
            
            conn.commit()
            print(f"✅ 成功保存 {saved} 条记录", flush=True)
            return jsonify({"code": 0, "message": "上传成功", "saved": saved})
        
        except Exception as e:
            return jsonify({"error": f"保存失败: {str(e)}"}), 500
        finally:
            conn.close()
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────
# /get_family_patient  家属查找绑定的患者
# ──────────────────────────────────────────────
@app.route("/get_family_patient", methods=["GET"])
def get_family_patient():
    viewer_id = request.args.get("viewerId", "").strip()
    if not viewer_id:
        return jsonify({"error": "缺少 viewerId 参数"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute("""
                SELECT patient_id, name, created_at
                FROM family_bindings
                WHERE family_id = %s AND status='active'
                ORDER BY created_at DESC
                LIMIT 1
            """, (viewer_id,))
        else:
            cursor.execute("""
                SELECT patient_id, name, created_at
                FROM family_bindings
                WHERE family_id = ? AND status='active'
                ORDER BY created_at DESC
                LIMIT 1
            """, (viewer_id,))
        row = cursor.fetchone()
        if row:
            name_val = row["name"] if "name" in row.keys() else (row["patient_id"] or "")
            return jsonify({
                "code": 0,
                "patientId": row["patient_id"],
                "patientName": name_val or row["patient_id"],
                "relation": name_val or "家人"
            })
        else:
            return jsonify({"code": 0, "patientId": "", "patientName": "", "message": "未找到绑定"})
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /get_binding_status  检查用户绑定状态 + 预警摘要（APP式启动接口）
# 前端每次启动/恢复时调用此接口，作为所有绑定数据的唯一真相源
# ──────────────────────────────────────────────
@app.route("/get_binding_status", methods=["GET"])
def get_binding_status():
    user_id = request.args.get("userId", "").strip()
    if not user_id:
        return jsonify({"error": "缺少 userId 参数"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        result = {
            "hasFamilyBinding": False,
            "hasDoctorBinding": False,
            "familyPatients": [],
            "doctorPatients": [],
            "familyAlertRisk": "none",
            "doctorAlertCount": 0,
            "familyAlertSummary": None,
            "doctorAlertSummary": None
        }

        # ── 1. 检查家属绑定（仅 active，支持多患者） ──
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT patient_id, name, created_at, status FROM family_bindings WHERE family_id=%s AND status='active' ORDER BY created_at DESC",
                (user_id,)
            )
        else:
            cursor.execute(
                "SELECT patient_id, name, created_at, status FROM family_bindings WHERE family_id=? AND status='active' ORDER BY created_at DESC",
                (user_id,)
            )
        family_rows = cursor.fetchall()
        if family_rows:
            result["hasFamilyBinding"] = True
            family_patients = []
            max_risk = "none"
            for row in family_rows:
                r = dict(row) if not isinstance(row, dict) else row
                pid = r.get("patient_id") or r["patient_id"]
                pname = r.get("name", "") or pid
                if USE_CLOUD_DB:
                    cursor.execute(
                        "SELECT sbp, dbp, hr, risk_level, datetime FROM measurements WHERE user_id=%s ORDER BY datetime DESC LIMIT 1",
                        (pid,)
                    )
                else:
                    cursor.execute(
                        "SELECT sbp, dbp, hr, risk_level, datetime FROM measurements WHERE user_id=? ORDER BY datetime DESC LIMIT 1",
                        (pid,)
                    )
                meas = cursor.fetchone()
                risk = "none"; sbp = None; dbp = None; hr = None; d = None
                if meas:
                    sbp = int(meas["sbp"] or 0)
                    dbp = int(meas["dbp"] or 0)
                    hr = int(meas["hr"] or 0)
                    rl = meas["risk_level"] if "risk_level" in meas.keys() else ""
                    d = str(meas["datetime"])[:10]
                    if rl == "high" or sbp >= 160 or dbp >= 100:
                        risk = "high"
                    elif rl == "moderate" or sbp >= 140 or dbp >= 90:
                        risk = "moderate"
                    elif sbp > 0:
                        risk = "normal"
                if risk == "high" or (risk == "moderate" and max_risk != "high"):
                    max_risk = risk
                family_patients.append({
                    "patientId": pid,
                    "patientName": pname,
                    "risk": risk,
                    "sbp": sbp,
                    "dbp": dbp,
                    "hr": hr,
                    "date": d
                })

            result["familyPatients"] = family_patients
            result["familyAlertRisk"] = max_risk
            fp = family_patients[0]
            result["familyAlertSummary"] = {
                "patientId": fp["patientId"],
                "patientName": fp["patientName"],
                "risk": fp["risk"],
                "sbp": fp["sbp"],
                "dbp": fp["dbp"],
                "date": fp["date"]
            }

        # ── 2. 检查医生绑定（仅 active，全量患者+风险） ──
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT patient_id, doctor_name, hospital, department, created_at, status FROM doctor_bindings WHERE doctor_id=%s AND status='active' ORDER BY created_at DESC",
                (user_id,)
            )
        else:
            cursor.execute(
                "SELECT patient_id, doctor_name, hospital, department, created_at, status FROM doctor_bindings WHERE doctor_id=? AND status='active' ORDER BY created_at DESC",
                (user_id,)
            )
        doctor_rows = cursor.fetchall()
        if doctor_rows:
            result["hasDoctorBinding"] = True
            doctor_patients = []
            high_count = moderate_count = normal_count = unmonitored_count = 0
            for row in doctor_rows:
                r = dict(row) if not isinstance(row, dict) else row
                pid = r.get("patient_id") or r["patient_id"]
                dname = r.get("doctor_name", "") or ""
                hosp = r.get("hospital", "") or ""
                dept = r.get("department", "") or ""
                if USE_CLOUD_DB:
                    cursor.execute(
                        "SELECT sbp, dbp, hr, risk_level, datetime FROM measurements WHERE user_id=%s ORDER BY datetime DESC LIMIT 1",
                        (pid,)
                    )
                else:
                    cursor.execute(
                        "SELECT sbp, dbp, hr, risk_level, datetime FROM measurements WHERE user_id=? ORDER BY datetime DESC LIMIT 1",
                        (pid,)
                    )
                m = cursor.fetchone()
                risk = "none"; sbp = None; dbp = None; hr = None; d = None
                if not m:
                    unmonitored_count += 1
                    risk = "none"
                else:
                    sbp = int(m["sbp"] or 0)
                    dbp = int(m["dbp"] or 0)
                    hr = int(m["hr"] or 0)
                    rl = m["risk_level"] if "risk_level" in m.keys() else ""
                    d = str(m["datetime"])[:10]
                    if rl == "high" or sbp >= 160 or dbp >= 100:
                        high_count += 1; risk = "high"
                    elif rl == "moderate" or sbp >= 140 or dbp >= 90:
                        moderate_count += 1; risk = "moderate"
                    elif sbp > 0:
                        normal_count += 1; risk = "normal"
                doctor_patients.append({
                    "patientId": pid,
                    "doctorName": dname,
                    "hospital": hosp,
                    "department": dept,
                    "risk": risk,
                    "sbp": sbp,
                    "dbp": dbp,
                    "hr": hr,
                    "date": d
                })

            result["doctorPatients"] = doctor_patients
            result["doctorAlertCount"] = high_count + moderate_count
            result["doctorAlertSummary"] = {
                "totalPatients": len(doctor_rows),
                "highRiskCount": high_count,
                "moderateRiskCount": moderate_count,
                "normalCount": normal_count,
                "unmonitoredCount": unmonitored_count
            }

        return jsonify({"code": 0, "data": result})
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /get_patients_risk_summary  医生查看患者风险汇总
# ──────────────────────────────────────────────
@app.route("/get_patients_risk_summary", methods=["GET"])
def get_patients_risk_summary():
    doctor_id = request.args.get("doctorId", "").strip()
    if not doctor_id:
        return jsonify({"error": "缺少 doctorId 参数"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute("SELECT patient_id FROM doctor_bindings WHERE doctor_id = %s AND status='active'", (doctor_id,))
        else:
            cursor.execute("SELECT patient_id FROM doctor_bindings WHERE doctor_id = ? AND status='active'", (doctor_id,))
        bindings = cursor.fetchall()
        patient_ids = [row["patient_id"] for row in bindings]

        if not patient_ids:
            return jsonify({"code": 0, "highRiskCount": 0, "moderateRiskCount": 0, "unmonitoredCount": 0, "details": {}})

        risk_detail = {}
        high_count = moderate_count = unmonitored_count = 0

        for pid in patient_ids:
            if USE_CLOUD_DB:
                cursor.execute("SELECT sbp, dbp, datetime, risk_level FROM measurements WHERE user_id = %s ORDER BY datetime DESC LIMIT 1", (pid,))
            else:
                cursor.execute("SELECT sbp, dbp, datetime, risk_level FROM measurements WHERE user_id = ? ORDER BY datetime DESC LIMIT 1", (pid,))
            row = cursor.fetchone()
            if not row:
                unmonitored_count += 1
                risk_detail[pid] = {"risk": "none", "sbp": None, "dbp": None, "date": None}
                continue

            sbp = int(row["sbp"] or 0)
            dbp = int(row["dbp"] or 0)
            rl = row["risk_level"] if "risk_level" in row.keys() else ""

            if rl == "high" or sbp >= 160 or dbp >= 100:
                high_count += 1
                risk_detail[pid] = {"risk": "high", "sbp": sbp, "dbp": dbp, "date": str(row["datetime"])[:10]}
            elif rl == "moderate" or sbp >= 140 or dbp >= 90:
                moderate_count += 1
                risk_detail[pid] = {"risk": "moderate", "sbp": sbp, "dbp": dbp, "date": str(row["datetime"])[:10]}
            else:
                risk_detail[pid] = {"risk": "normal", "sbp": sbp, "dbp": dbp, "date": str(row["datetime"])[:10]}

        return jsonify({
            "code": 0,
            "highRiskCount": high_count,
            "moderateRiskCount": moderate_count,
            "unmonitoredCount": unmonitored_count,
            "details": risk_detail
        })
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 80))
    app.run(host="0.0.0.0", port=port, debug=False)
