# -*- coding: utf-8 -*-
# binding_views.py
from flask import Blueprint, jsonify, request
import database
import json
import random
import string
import time
from datetime import datetime
from engine.cardiovascular_engine import CardiovascularEngine
from services import save_measurement

binding_bp = Blueprint('binding', __name__)


def _generate_user_id(role):
    """
    ★ 新增：user_id 生成逻辑从前端搬到后端——
    身份识别体系重构后，openid 才是真正的身份锚点，user_id 只是内部主键，
    由后端统一决定何时创建新的（不再信任前端生成的随机字符串）。
    """
    prefix = {'doctor': 'D', 'family': 'F'}.get(role, 'U')
    ts = format(int(time.time() * 1000), 'x').upper()
    rand = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"{prefix}{ts}{rand}"


def _resolve_viewer_user_id(cursor, ph, openid, viewer_name, role):
    """
    绑定流程里，家属/医生本人可能还没有完整走过 register_user 流程，
    这里用 openid 兜底建立一条基础记录（只有姓名和角色）。
    和 register_user 一样，只认 openid，不依赖前端传的 user_id——
    这样即使这个人清了本地缓存重新点开邀请链接，也会被识别成同一个人，
    而不是被当成新用户重复建一条记录。
    """
    cursor.execute(f"SELECT user_id, name FROM users WHERE openid = {ph}", (openid,))
    existing = cursor.fetchone()
    if existing:
        existing = dict(existing)
        # ★ 修复：已存在的账号，之前这里直接返回 user_id，从不更新 users.name，
        #   导致早期版本(改成"必须手填真实姓名"之前)自动生成的描述文字/
        #   用户填错的名字会一直冻结在数据库里，即使这次重新确认绑定填了
        #   正确姓名也不会生效——留言列表里 sender_name 每次都现查 users.name，
        #   查到的一直是那条陈旧数据。现在改成：这次填的姓名如果跟已存的不一样，
        #   就更新过去(以本人最近一次填写的为准，姓名不该因为角色不同而不同)。
        if viewer_name and viewer_name != existing.get('name'):
            cursor.execute(
                f"UPDATE users SET name = {ph} WHERE user_id = {ph}",
                (viewer_name, existing['user_id'])
            )
        return existing['user_id']
    new_id = _generate_user_id(role)
    cursor.execute(
        f"INSERT INTO users (user_id, name, role, openid) VALUES ({ph}, {ph}, {ph}, {ph})",
        (new_id, viewer_name, role, openid)
    )
    return new_id

def parse_to_datetime_obj(ts):
    if isinstance(ts, datetime): return ts
    if isinstance(ts, str):
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M"):
            try: return datetime.strptime(ts, fmt)
            except ValueError: continue
    return datetime.now()

@binding_bp.route('/get_binding_status', methods=['GET'])
def get_binding_status():
    """对应前端 app.js 启动和轮询调用"""
    user_id = request.args.get('userId') or request.args.get('user_id')
    if not user_id:
        return jsonify({"code": 1, "msg": "Missing userId"}), 400

    conn = database.get_db()
    cursor = conn.cursor()
    ph = database.get_placeholder()

    has_family_binding = False
    family_patients = []
    family_alert_risk = "none"
    family_patient_name = ""

    has_doctor_binding = False
    doctor_patients = []
    doctor_alert_count = 0

    try:
        # 自适应探测 family_bindings 表结构，100% 防止 no such column
        columns = database.get_table_columns(cursor, 'family_bindings')

        family_bonds = []
        if columns:
            target_col = "sharer_id" if "sharer_id" in columns else ("user_id" if "user_id" in columns else columns[1])
            # ★ 改：加上 status='active' 过滤，否则取消绑定(status改成cancelled)后
            #   这里依然会把已取消的关系当成有效绑定返回，取消功能会形同虚设
            cursor.execute(f"SELECT * FROM family_bindings WHERE {target_col} = {ph} AND status = 'active'", (user_id,))
            family_bonds = cursor.fetchall()

        if family_bonds:
            has_family_binding = True
            p_id_holder = family_bonds[0]["patient_id"] if "patient_id" in columns else user_id

            # ★ 新增：查患者的 name + birth_date，拼成"姓名(出生日期)"展示名，
            #   方便家属一眼认出这是谁，而不是只看到一串内部 user_id
            cursor.execute(f"SELECT name, birth_date FROM users WHERE user_id = {ph}", (p_id_holder,))
            p_user = cursor.fetchone()
            if p_user and p_user.get('name'):
                bdate = p_user.get('birth_date') or ''
                family_patient_name = f"{p_user['name']}({bdate})" if bdate else p_user['name']
            else:
                family_patient_name = "我的关联被监护家人"  # 兜底：查不到患者资料时的老文案

            # ★ 改：先算一次未读数，避免同一段查询被调用两次
            family_unread_count = _count_unread_feedback(cursor, ph, user_id, p_id_holder)
            family_patients.append({
                "patientId": p_id_holder,
                "patientName": family_patient_name,
                "hasUnreadFeedback": family_unread_count > 0,
                "unreadFeedbackCount": family_unread_count
            })
            # 💡 动态健康监控：抽取时间序列为家属端运算当前警报级别
            family_alert_risk = _calculate_realtime_risk(conn, cursor, p_id_holder)

        # 检查医生监护池
        # ★ 改：加上 status='active' 过滤，同上，否则取消绑定后依然显示
        cursor.execute(f"SELECT * FROM doctor_bindings WHERE doctor_id = {ph} AND status = 'active'", (user_id,))
        doctor_bonds = cursor.fetchall()

        if doctor_bonds:
            has_doctor_binding = True
            for bond in doctor_bonds:
                p_id = bond['patient_id']
                # ★ 改：同时查 age、gender——患者列表页需要姓名/年龄/性别分开展示，
                #   不能只用拼好的 patientName（那是"姓名(出生日期)"格式，给仪表盘头部用的）
                cursor.execute(f"SELECT name, birth_date, age, gender FROM users WHERE user_id = {ph}", (p_id,))
                p_user = cursor.fetchone()
                if p_user and p_user.get('name'):
                    bdate = p_user.get('birth_date') or ''
                    p_name = f"{p_user['name']}({bdate})" if bdate else p_user['name']
                    raw_name = p_user['name']
                    age = p_user.get('age')
                    gender = p_user.get('gender') or ''
                else:
                    p_name = "签约患者"  # 兜底：查不到患者资料时的老文案
                    raw_name = "签约患者"
                    age = None
                    gender = ''

                # ★ 改：把风险等级也一起返回，供医生端列表展示状态徽章 + 排序依据
                risk = _calculate_realtime_risk(conn, cursor, p_id)
                doctor_unread_count = _count_unread_feedback(cursor, ph, user_id, p_id, doctor_id=user_id)
                doctor_patients.append({
                    "patientId": p_id,
                    "patientName": p_name,
                    "name": raw_name,
                    "age": age,
                    "gender": gender,
                    "riskLevel": risk,
                    "hasUnreadFeedback": doctor_unread_count > 0,
                    "unreadFeedbackCount": doctor_unread_count
                })
                # 如果医生监护的某位患者触发了高风险，医生的未读红点加1
                if risk in ["high", "critical"]:
                    doctor_alert_count += 1

            # ★ 新增：按病情严重程度排序，最需要关注的患者排在列表最前面（预警优先展示）
            #   闭环理念：医生端不是"有病来找",而是持续追踪，列表顺序本身就是一种预警
            risk_priority = {"critical": 4, "high": 3, "moderate": 2, "low": 1, "none": 0}
            doctor_patients.sort(key=lambda p: risk_priority.get(p["riskLevel"], 0), reverse=True)

        # ★ 新增：这个用户"作为患者"，自己名下所有医生线的未读留言总数——
        #   之前这里只统计了"作为家属/医生查看别人时"的未读(上面 family_unread_count/
        #   doctor_unread_count 那两处)，从没统计过"作为患者，别的医生给我发的留言
        #   我还没看"这种情况，导致 more 页面"健康反馈"入口一直没有未读角标。
        #   跟角色无关，任何人都可能同时"作为患者"被医生绑定，所以放在 family/doctor
        #   两个 if 分支外面，始终统计。
        self_unread_count = _count_unread_feedback(cursor, ph, user_id, user_id)

        return jsonify({
            "code": 0,
            "msg": "success",
            "data": {
                "hasFamilyBinding": has_family_binding,
                "familyPatients": family_patients,
                "hasDoctorBinding": has_doctor_binding,
                "doctorPatients": doctor_patients,
                "familyAlertRisk": family_alert_risk,
                "doctorAlertCount": doctor_alert_count,
                "familyAlertSummary": {"patientName": family_patient_name},
                "selfUnreadFeedbackCount": self_unread_count
            }
        })

    except Exception as e:
        print(f"⚠️ [Binding status proxy auto sync active]: {e}")
        return jsonify({
            "code": 0, "msg": "degraded success",
            "data": {
                "hasFamilyBinding": False, "familyPatients": [],
                "hasDoctorBinding": False, "doctorPatients": [],
                "familyAlertRisk": "none", "doctorAlertCount": 0, "familyAlertSummary": {"patientName": ""},
                "selfUnreadFeedbackCount": 0
            }
        })

def _count_unread_feedback(cursor, ph, viewer_id, patient_id, doctor_id=None):
    """
    统计某个查看者(viewer_id)对某个患者(patient_id)有几条未读留言。
    ★ 改：留言线现在按"基础线(患者+家属专属)+每个医生一条诊疗线"拆分，
    未读计数也要按线区分：
    - 传 doctor_id(可以是空字符串''，代表基础线；或者具体医生 id，代表诊疗线)：
      只统计这一条线
    - 不传 doctor_id(None)：统计"基础线"+这个患者名下所有 active 医生诊疗线
      的未读总和(家属/患者看汇总角标时用这个，因为他们能看所有线，需要一个
      加总的数字；医生只能看自己那条诊疗线，永远传具体 doctor_id，不会走进
      这个汇总分支，所以汇总里包含基础线不会影响医生视角)
    "未读"指：不是自己发的、且发送时间晚于自己上次查看这条线的记录
    （或者从没查看过、但确实有别人发过的留言，这时全部算未读）。
    各端(患者/家属/医生)、各条线的已读进度分别记在 feedback_read_progress 表里，互不影响。
    """
    try:
        if doctor_id is not None:
            cursor.execute(
                f"SELECT last_read_at FROM feedback_read_progress "
                f"WHERE viewer_id={ph} AND patient_id={ph} AND doctor_id={ph}",
                (viewer_id, patient_id, doctor_id)
            )
            progress = cursor.fetchone()
            last_read_at = dict(progress)['last_read_at'] if progress else None

            if last_read_at:
                cursor.execute(
                    f"SELECT COUNT(*) AS cnt FROM feedbacks "
                    f"WHERE to_id={ph} AND doctor_id={ph} AND from_id != {ph} AND created_at > {ph}",
                    (patient_id, doctor_id, viewer_id, last_read_at)
                )
            else:
                cursor.execute(
                    f"SELECT COUNT(*) AS cnt FROM feedbacks WHERE to_id={ph} AND doctor_id={ph} AND from_id != {ph}",
                    (patient_id, doctor_id, viewer_id)
                )
            row = cursor.fetchone()
            return dict(row)['cnt'] if row else 0

        # 不传 doctor_id：汇总"基础线"+这个患者名下所有active医生线的未读总和
        total = _count_unread_feedback(cursor, ph, viewer_id, patient_id, doctor_id='')  # 基础线
        cursor.execute(
            f"SELECT doctor_id FROM doctor_bindings WHERE patient_id={ph} AND status='active'",
            (patient_id,)
        )
        doctor_ids = [dict(r)['doctor_id'] for r in cursor.fetchall()]
        for did in doctor_ids:
            total += _count_unread_feedback(cursor, ph, viewer_id, patient_id, doctor_id=did)
        return total
    except Exception:
        # 已读状态查询失败不应该影响整个患者列表的加载，安静地当作"没有未读"处理
        return 0


def _calculate_realtime_risk(conn, cursor, user_id):
    """辅助长效健康闭环：时序数据清洗并唤醒核心引擎"""
    ph = database.get_placeholder()
    try:
        cursor.execute(
            f"SELECT sbp, dbp, hr, symptoms, datetime FROM measurements WHERE user_id = {ph} ORDER BY datetime DESC LIMIT 20",
            (user_id,)
        )
        rows = cursor.fetchall()
        if not rows: return "none"

        records = []
        for r in rows:
            d_r = dict(r)
            try:
                if isinstance(d_r.get('symptoms'), str):
                    d_r['symptoms'] = json.loads(d_r['symptoms'] or '[]')
            except: d_r['symptoms'] = []
            d_r['datetime'] = parse_to_datetime_obj(d_r.get('datetime'))
            records.append(d_r)

        if len(records) < 2:
            return records[0].get("risk_level", "low")

        engine = CardiovascularEngine(history=records[1:], current=records[0])
        res = engine.run_all_diagnostics()
        return res.get("risk_level", "low")
    except:
        return "low"

@binding_bp.route('/register_user', methods=['POST'])
def register_user():
    """
    真正的核心大脑：将前端皮囊上报的用户数据，写入数据库。
    ★ 改：身份识别体系重构——openid 才是这个人的真正身份锚点，
       user_id 现在由后端根据 openid 决定：老用户（openid 已存在）
       复用原有 user_id，新用户由后端生成一个新的。
       不再由前端随机生成 user_id 再传上来——这样即使前端本地存储丢失
       导致生成了不同的临时标识，只要 openid 一致，后端总能找到同一条
       用户记录，不会产生重复账号，也不会造成"清缓存=失联"的问题。
    """
    try:
        data = request.json or {}
        openid = data.get('openid')
        name = data.get('name')
        age = data.get('age', 0)
        gender = data.get('gender', '')
        role = data.get('role', 'user')
        birth_date = data.get('birth_date') or data.get('birthDate') or ''
        # ★ 新增：既往病史——之前前端onSubmit压根没把这个字段发过来，这里
        #   一直没读取。用JSON字符串存(跟history_views.py存symptoms字段是
        #   同一种做法)，读取时json.loads解析回数组。
        health_history = data.get('health_history') or data.get('healthHistory') or []
        health_history_json = json.dumps(health_history, ensure_ascii=False)

        if not openid or not name:
            return jsonify({"code": -1, "msg": "核心字段缺失（缺少 openid 或姓名），大脑拒绝写入"}), 400

        conn = database.get_db()
        cursor = conn.cursor()
        ph = database.get_placeholder()

        # ★ 改：按 openid 查找，而不是按前端传来的 user_id 查找
        cursor.execute(f"SELECT user_id FROM users WHERE openid = {ph}", (openid,))
        existing = cursor.fetchone()

        if existing:
            existing = dict(existing)
            user_id = existing['user_id']
            cursor.execute(f"""
                UPDATE users SET name={ph}, age={ph}, gender={ph}, role={ph}, birth_date={ph}, health_history={ph} WHERE openid={ph}
            """, (name, age, gender, role, birth_date, health_history_json, openid))
        else:
            user_id = _generate_user_id(role)
            cursor.execute(f"""
                INSERT INTO users (user_id, name, age, gender, role, birth_date, openid, health_history) 
                VALUES ({ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph})
            """, (user_id, name, age, gender, role, birth_date, openid, health_history_json))

        conn.commit()
        print(f"🟢 [DB] 用户 {name}({user_id}) 注册数据成功写入！openid={openid}", flush=True)

        return jsonify({
            "code": 0,
            "msg": "注册成功",
            "data": {"userId": user_id}
        }), 200

    except Exception as e:
        print(f"❌ [DB] 注册存盘发生坍塌: {str(e)}", flush=True)
        return jsonify({"code": 500, "msg": f"大脑数据库写入异常: {str(e)}"}), 500


# ============================================================
# ★ 新增：V10 邀请绑定主链路（永久链接模式）
# 前端 intro.js 里调用的 cloudService.validateInvite / bindByInvite
# 对应的后端实现，此前完全缺失，导致邀请绑定功能实际不可用。
# 分享链接里携带的是患者的 user_id 本身（永久有效），不是一次性 token。
# ============================================================

@binding_bp.route('/validate_invite', methods=['POST'])
def validate_invite():
    """
    邀请确认页打开时调用：校验链接里的 patientId 是否对应真实存在的患者，
    并返回患者展示名 + 一份简单摘要，供家属/医生确认"这是不是我要绑定的人"。
    """
    data = request.json or {}
    patient_id = data.get('patientId')
    role = data.get('role')  # 'family' | 'doctor'，当前校验逻辑不区分角色，仅透传保留

    if not patient_id:
        return jsonify({"code": 1, "msg": "缺少 patientId"}), 400

    conn = database.get_db()
    cursor = conn.cursor()
    ph = database.get_placeholder()

    try:
        cursor.execute(f"SELECT name, birth_date FROM users WHERE user_id = {ph}", (patient_id,))
        patient = cursor.fetchone()
        # ★ 改：加 dict() 转换——sqlite3.Row 不支持 .get()，MySQL 的 DictCursor 结果 dict() 是幂等操作，
        #   之前这里漏了这一步，导致一旦数据库意外降级成 SQLite 模式，这个接口就会直接报错，
        #   而报错又被 except 兜底成"患者不存在"，掩盖了真实原因
        patient = dict(patient) if patient else None

        if not patient or not patient.get('name'):
            return jsonify({"code": -1, "msg": "邀请链接无效或患者不存在"}), 404

        bdate = patient.get('birth_date') or ''
        patient_name = f"{patient['name']}({bdate})" if bdate else patient['name']

        # 患者摘要：取最近一次测量记录，给确认页一个基本的"认人"参考
        cursor.execute(
            f"SELECT sbp, dbp, risk_level, datetime FROM measurements WHERE user_id = {ph} ORDER BY datetime DESC LIMIT 1",
            (patient_id,)
        )
        latest = cursor.fetchone()
        latest = dict(latest) if latest else None
        patient_summary = {
            "name": patient_name,
            "latestSbp": latest['sbp'] if latest else None,
            "latestDbp": latest['dbp'] if latest else None,
            "latestRiskLevel": latest['risk_level'] if latest else None,
            "latestDatetime": latest['datetime'] if latest else None,
        }

        return jsonify({
            "code": 0,
            "msg": "success",
            "data": {
                "patientName": patient_name,
                "patientSummary": patient_summary
            }
        })

    except Exception as e:
        print(f"❌ [validate_invite] 校验失败: {e}", flush=True)
        return jsonify({"code": -1, "msg": f"校验失败: {str(e)}"}), 500


@binding_bp.route('/bind_by_invite', methods=['POST'])
def bind_by_invite():
    """
    用户点击"确认绑定"后调用：写入 family_bindings 或 doctor_bindings。
    如果之前已经绑定过同一对 (viewer, patient)，则把状态刷新为 active，
    不会因为 UNIQUE 约束报错而失败（幂等处理，允许重复点击/重新绑定）。

    ★ 改：身份识别体系重构——前端不再传 viewerId（那是前端本地生成的、
       清缓存就会变的临时标识），改为传 viewerOpenid，后端用 openid
       查找/创建这个人的真正 user_id（见 _resolve_viewer_user_id）。
       这样即使这个家属/医生之前清过缓存、本地记不清自己是谁，
       只要还是同一个微信号，这次绑定依然会落在同一条用户记录和
       同一段绑定关系上，不会产生"查无此人"或重复账号的问题。
       返回结果里带上 viewerId，前端要用这个值覆盖本地存储。
    """
    data = request.json or {}
    patient_id = data.get('patientId')
    role = data.get('role')
    viewer_openid = data.get('viewerOpenid') or data.get('openid')
    viewer_name = data.get('viewerName') or ''
    hospital = data.get('hospital') or ''
    department = data.get('department') or ''

    if not patient_id or not role or not viewer_openid:
        return jsonify({"code": -1, "msg": "缺少必要参数"}), 400
    if role not in ('family', 'doctor'):
        return jsonify({"code": -1, "msg": "role 参数不合法"}), 400

    conn = database.get_db()
    cursor = conn.cursor()
    ph = database.get_placeholder()

    try:
        # 再次确认患者存在（防止确认页打开后患者数据被删除的边界情况）
        # ★ 改：同时查出患者自己的 openid，用来在下面提前判断"是不是自己绑自己"，
        #   必须在调用 _resolve_viewer_user_id 之前就查出来并比对完——那个函数
        #   一旦被调用，只要 openid 命中已存在账号，就会顺手更新 users.name，
        #   如果等它跑完之后才发现是自己绑自己，姓名覆盖这个副作用已经发生了，
        #   校验就晚了，等于白拦
        cursor.execute(f"SELECT name, openid FROM users WHERE user_id = {ph}", (patient_id,))
        patient = cursor.fetchone()
        if not patient:
            return jsonify({"code": -1, "msg": "患者不存在，绑定失败"}), 404
        patient = dict(patient)
        patient_name = patient.get('name') or ''

        # ★ 新增：禁止绑定自己——在真正解析/写入 viewer 身份之前就先比对 openid
        #   拦下来。这不是理论上的边界情况：反复出现过的"姓名显示错乱/角色判断
        #   混乱"，根源都是同一个真实账号用来"自己绑自己"做测试，导致 users 表
        #   里同一行记录被家属/医生/患者三种身份的姓名反复互相覆盖——一个人
        #   不可能同时是自己的家属或医生，数据结构上就不成立，必须在这里堵死。
        if viewer_openid and viewer_openid == patient.get('openid'):
            return jsonify({"code": -1, "msg": "不能绑定自己，请使用其他微信账号确认绑定"}), 400

        # ★ 改：用 openid 找/建家属或医生本人的真正 user_id
        viewer_id = _resolve_viewer_user_id(cursor, ph, viewer_openid, viewer_name, role)

        # ★ 双保险：万一上面 openid 比对因为边界情况漏判(比如患者账号
        #   openid 字段本身是空的)，这里再用解析出来的 viewer_id 比对一次，
        #   避免真的写出一条自己绑自己的关系记录
        if viewer_id == patient_id:
            return jsonify({"code": -1, "msg": "不能绑定自己，请使用其他微信账号确认绑定"}), 400

        if role == 'family':
            cursor.execute(
                f"SELECT id FROM family_bindings WHERE family_id={ph} AND patient_id={ph}",
                (viewer_id, patient_id)
            )
            existing = cursor.fetchone()
            if existing:
                # ★ 修复：这里之前存的是 patient_name(患者自己的名字)，导致
                #   family_bindings.name 不管绑定的是哪个家属，值永远等于患者自己的
                #   名字——get_my_bindings 返回给"绑定管理"页面时，所有家属条目
                #   就都显示成了患者本人的名字。改成存 viewer_name(这个家属自己的名字)。
                cursor.execute(
                    f"UPDATE family_bindings SET status='active', name={ph} WHERE family_id={ph} AND patient_id={ph}",
                    (viewer_name, viewer_id, patient_id)
                )
            else:
                cursor.execute(
                    f"INSERT INTO family_bindings (family_id, patient_id, name, status) VALUES ({ph}, {ph}, {ph}, 'active')",
                    (viewer_id, patient_id, viewer_name)
                )
        else:  # doctor
            cursor.execute(
                f"SELECT id FROM doctor_bindings WHERE doctor_id={ph} AND patient_id={ph}",
                (viewer_id, patient_id)
            )
            existing = cursor.fetchone()
            if existing:
                cursor.execute(
                    f"""UPDATE doctor_bindings SET status='active', doctor_name={ph}, hospital={ph}, department={ph}
                        WHERE doctor_id={ph} AND patient_id={ph}""",
                    (viewer_name, hospital, department, viewer_id, patient_id)
                )
            else:
                cursor.execute(
                    f"""INSERT INTO doctor_bindings (doctor_id, patient_id, doctor_name, hospital, department, status)
                        VALUES ({ph}, {ph}, {ph}, {ph}, {ph}, 'active')""",
                    (viewer_id, patient_id, viewer_name, hospital, department)
                )

        conn.commit()
        print(f"🟢 [DB] 绑定成功: {role} {viewer_id} -> patient {patient_id}", flush=True)

        return jsonify({
            "code": 0,
            "msg": "绑定成功",
            "data": {"viewerId": viewer_id}
        })

    except Exception as e:
        print(f"❌ [bind_by_invite] 绑定失败: {e}", flush=True)
        return jsonify({"code": -1, "msg": f"绑定失败: {str(e)}"}), 500


# ============================================================
# ★ 新增：医生端患者列表分页搜索（幽灵接口第二批）
# 前端 doctor/dashboard/dashboard.js 的 loadBoundPatients() 一直在调用这个接口，
# 之前后端完全没实现，请求会静默失败。主体患者列表走 get_binding_status 不受影响，
# 这个接口只补齐"下拉框内按姓名搜索 + 翻页"这个辅助功能。
# ============================================================

@binding_bp.route('/get_doctor_patients', methods=['GET'])
def get_doctor_patients():
    """
    医生端患者列表：支持按姓名搜索、按风险等级降序排序、分页。
    患者数量级别不大，这里先查出该医生的全部绑定患者，
    在内存里完成排序/搜索/分页，不做数据库层面的分页优化。
    """
    doctor_id = request.args.get('doctorId')
    if not doctor_id:
        return jsonify({"code": 1, "msg": "缺少 doctorId"}), 400

    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('pageSize', 20))
    except ValueError:
        page, page_size = 1, 20
    keyword = (request.args.get('keyword') or '').strip()

    conn = database.get_db()
    cursor = conn.cursor()
    ph = database.get_placeholder()

    try:
        # ★ 改：同样加上 status='active' 过滤
        cursor.execute(f"SELECT * FROM doctor_bindings WHERE doctor_id = {ph} AND status = 'active'", (doctor_id,))
        doctor_bonds = cursor.fetchall()

        all_patients = []
        for bond in doctor_bonds:
            bond = dict(bond)  # ★ 统一转 dict：sqlite3.Row 不支持 .get()，MySQL 的 dict 结果 dict() 是幂等操作，两边都安全
            p_id = bond['patient_id']
            cursor.execute(f"SELECT name, birth_date, age, gender FROM users WHERE user_id = {ph}", (p_id,))
            p_user = cursor.fetchone()
            p_user = dict(p_user) if p_user else None

            raw_name = p_user['name'] if (p_user and p_user.get('name')) else ''
            # ★ 关键词搜索只匹配姓名，不匹配内部 user_id
            #   （患者内部ID不该作为医生可用的查找方式，姓名同名概率低，够用）
            if keyword and keyword not in raw_name:
                continue

            bdate = p_user.get('birth_date') if p_user else ''
            p_name = f"{raw_name}({bdate})" if (raw_name and bdate) else (raw_name or '签约患者')
            age = p_user.get('age') if p_user else None
            gender = (p_user.get('gender') or '') if p_user else ''

            risk = _calculate_realtime_risk(conn, cursor, p_id)
            all_patients.append({
                "patientId": p_id,
                "patientName": p_name,
                "name": raw_name or '签约患者',
                "age": age,
                "gender": gender,
                "riskLevel": risk,
                "doctorName": bond.get('doctor_name') or '',
                "hospital": bond.get('hospital') or '',
                "department": bond.get('department') or ''
            })

        # ★ 按风险等级降序排序，和 get_binding_status 保持一致的优先级
        risk_priority = {"critical": 4, "high": 3, "moderate": 2, "low": 1, "none": 0}
        all_patients.sort(key=lambda p: risk_priority.get(p["riskLevel"], 0), reverse=True)

        total = len(all_patients)
        start = (page - 1) * page_size
        end = start + page_size
        page_data = all_patients[start:end]
        has_more = end < total

        return jsonify({
            "code": 0,
            "msg": "success",
            "data": page_data,
            "total": total,
            "hasMore": has_more
        })

    except Exception as e:
        print(f"❌ [get_doctor_patients] 查询失败: {e}", flush=True)
        return jsonify({"code": -1, "msg": f"查询失败: {str(e)}"}), 500


# ============================================================
# ★ 新增：历史数据批量上传（幽灵接口第二批，最后一个）
# ★ 改：原本设计走 wx.uploadFile 直连裸域名（不带 /api 前缀），实测在本项目
#   环境下从未跑通（404）。现改为前端把文件读成 base64、通过已验证可靠的
#   wx.cloud.callContainer 以普通 JSON 方式发送，因此这里和其他接口一样
#   走标准 /api 前缀，注册在 binding_bp 蓝图下，不再需要 app.py 里的特例路由。
#
# 支持 .xlsx/.xls（openpyxl 解析）和 .csv（Python 内置 csv 模块解析）。
# 必填列：日期、时间、收缩压、舒张压；心率可为空（默认75）。
# 脉压差由后端自动计算，不读取任何"脉压差"列。不解析"备注"列。
# 不触发分析引擎，纯批量存库——患者之后正常测量提交时，
# /api/analyze 会自动把全部历史（含这批导入的）一起纳入稳态计算。
# ============================================================

def _process_import_rows(rows, user_id, conn, cursor, ph):
    """
    统一处理逻辑：rows 是一个可迭代对象，每个元素是 dict（列名 -> 值），
    xlsx 和 csv 两条解析路径最终都转换成这种统一格式后传进来。
    返回 (imported, skipped)。

    ★ 改：之前这里是自己把"日期"+"时间"两列原始文本直接拼接成字符串
    （datetime_str = f"{date_val} {time_val}"），完全没有做任何格式清洗，
    直接原样 INSERT 进数据库。Excel 里用户手填的日期五花八门（比如不补零的
    "2026/3/9"），跟正常测量提交走的格式（"2026-08-03 10:59"）不一致时，
    会导致 measurements 表里同一列混进两种字符串格式——由于 get_history
    的 ORDER BY datetime 是纯字符串比较，不是真正按时间排序，这批格式不一致
    的批量导入数据会把真实的最新记录"挤"出查询结果之外（这正是"家属端/
    医生端看不到最新数据"这个问题的真正原因）。
    现在改成调用 services.py 里 save_measurement()/_format_record_for_db()
    这个所有存库路径共用的清洗函数，保证不管数据从哪条路径进来，
    datetime 字段最终都是统一的 "YYYY-MM-DD HH:MM:SS" 格式。
    """
    imported = 0
    skipped = 0

    for row in rows:
        date_val = (row.get('日期') or '').strip() if row.get('日期') else ''
        time_val = (row.get('时间') or '').strip() if row.get('时间') else ''
        sbp_val = (row.get('收缩压') or '').strip() if row.get('收缩压') else ''
        dbp_val = (row.get('舒张压') or '').strip() if row.get('舒张压') else ''
        hr_val = (row.get('心率') or '').strip() if row.get('心率') else ''

        if not date_val or not time_val or not sbp_val or not dbp_val:
            skipped += 1
            continue

        try:
            sbp = int(float(sbp_val))
            dbp = int(float(dbp_val))
        except (ValueError, TypeError):
            skipped += 1
            continue

        try:
            hr = int(float(hr_val)) if hr_val else 75
        except (ValueError, TypeError):
            hr = 75

        try:
            save_measurement(conn, cursor, {
                "user_id": user_id,
                "sbp": sbp,
                "dbp": dbp,
                "hr": hr,
                "datetime": f"{date_val} {time_val}"
            })
            imported += 1
        except Exception as e:
            print(f"⚠️ [upload_excel] 跳过一行导入失败: {e}", flush=True)
            skipped += 1

    return imported, skipped


@binding_bp.route('/upload_excel', methods=['POST'])
def upload_excel():
    """
    接收前端以 base64 编码传来的 Excel/CSV 文件内容，解析后批量导入 measurements 表。
    请求体（JSON）: { fileName, userId, fileBase64 }
    """
    import base64
    import io

    data = request.json or {}
    file_name = (data.get('fileName') or '').lower()
    user_id = data.get('userId') or data.get('user_id')
    file_base64 = data.get('fileBase64')

    if not file_base64:
        return jsonify({"code": -1, "msg": "未收到文件内容"}), 400
    if not user_id:
        return jsonify({"code": -1, "msg": "缺少 userId，无法确定归属患者"}), 400

    is_csv = file_name.endswith('.csv')
    is_excel = file_name.endswith('.xlsx') or file_name.endswith('.xls')
    if not is_csv and not is_excel:
        return jsonify({"code": -1, "msg": "仅支持 .xlsx/.xls/.csv 文件"}), 400

    try:
        file_bytes = base64.b64decode(file_base64)
    except Exception as e:
        return jsonify({"code": -1, "msg": f"文件内容解码失败: {str(e)}"}), 400

    rows = []
    required_cols = ["日期", "时间", "收缩压", "舒张压"]

    try:
        if is_excel:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = wb.active

            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                return jsonify({"code": -1, "msg": "文件为空或缺少表头"}), 400

            headers = [str(h).strip() if h else '' for h in header_row]
            missing_cols = [c for c in required_cols if c not in headers]
            if missing_cols:
                return jsonify({"code": -1, "msg": f"表头缺少必填列: {', '.join(missing_cols)}"}), 400

            for data_row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for idx, col_name in enumerate(headers):
                    if not col_name or idx >= len(data_row):
                        continue
                    val = data_row[idx]
                    row_dict[col_name] = str(val).strip() if val is not None else ''
                rows.append(row_dict)

        else:  # csv
            import csv
            try:
                text = file_bytes.decode('utf-8-sig')
            except UnicodeDecodeError:
                text = file_bytes.decode('gbk', errors='ignore')

            reader = csv.DictReader(io.StringIO(text))
            if not reader.fieldnames:
                return jsonify({"code": -1, "msg": "文件为空或缺少表头"}), 400

            headers = [h.strip() if h else '' for h in reader.fieldnames]
            missing_cols = [c for c in required_cols if c not in headers]
            if missing_cols:
                return jsonify({"code": -1, "msg": f"表头缺少必填列: {', '.join(missing_cols)}"}), 400

            for raw_row in reader:
                row_dict = {(k.strip() if k else k): (v.strip() if v else v) for k, v in raw_row.items()}
                rows.append(row_dict)

    except Exception as e:
        return jsonify({"code": -1, "msg": f"文件解析失败，请确认文件格式正确: {str(e)}"}), 400

    conn = database.get_db()
    cursor = conn.cursor()
    ph = database.get_placeholder()

    try:
        imported, skipped = _process_import_rows(rows, user_id, conn, cursor, ph)
        conn.commit()
        print(f"🟢 [DB] 批量导入完成: user={user_id}, 成功{imported}条, 跳过{skipped}条", flush=True)

        return jsonify({
            "code": 0,
            "msg": "导入完成",
            "data": {"imported": imported, "skipped": skipped}
        })

    except Exception as e:
        print(f"❌ [upload_excel] 导入失败: {e}", flush=True)
        return jsonify({"code": -1, "msg": f"导入失败: {str(e)}"}), 500


# ============================================================
# ★ 新增：微信身份确认（身份识别体系重构）
# ★ 改：最初设计是 wx.login() 拿 code → 后端调 api.weixin.qq.com 换 openid，
#   实测云托管容器出网访问该接口会撞上 SSL 自签名证书校验失败
#   （云托管出口网络对 HTTPS 出站流量做了中间代理，证书链和 requests 库
#   默认信任的证书链对不上）。经查微信官方文档确认：这套流程本来就是多余的——
#   小程序通过 wx.cloud.callContainer 调用云托管服务时，请求头会自动携带
#   X-WX-OPENID，不需要 wx.login+code 置换这一步。现在直接读请求头，
#   更简单、更安全（不再需要 AppSecret），也没有出网请求，不会有这个问题。
#
# 用途：小程序每次启动都调用一次，用请求头里的 openid 去数据库查这个人是谁。
# 这样即使本地存储被清空（清缓存/换设备），只要还是同一个微信号，
# 重新登录后依然能自动恢复身份和绑定关系，不用患者重新发邀请。
#
# 返回:
#   已注册用户 → { isNewUser: false, openid, userId, name, age, gender, role, birthDate }
#   新用户     → { isNewUser: true, openid }（前端拿着这个 openid 走注册/绑定流程）
# ============================================================

@binding_bp.route('/wx_login', methods=['POST'])
def wx_login():
    # ★ 改：openid 从请求头读取，微信云托管在 callContainer 调用时自动注入，
    #   请求头名不区分大小写，两种写法都尝试一下
    openid = request.headers.get('X-WX-OPENID') or request.headers.get('x-wx-openid')

    if not openid:
        print("❌ [wx_login] 请求头缺少 X-WX-OPENID，可能不是通过 wx.cloud.callContainer 调用的", flush=True)
        return jsonify({"code": -1, "msg": "未能获取微信身份信息，请通过小程序端调用"}), 400

    conn = database.get_db()
    cursor = conn.cursor()
    ph = database.get_placeholder()

    try:
        cursor.execute(
            f"SELECT user_id, name, age, gender, role, birth_date FROM users WHERE openid = {ph}",
            (openid,)
        )
        existing = cursor.fetchone()

        if existing:
            existing = dict(existing)
            return jsonify({
                "code": 0,
                "msg": "success",
                "data": {
                    "isNewUser": False,
                    "openid": openid,
                    "userId": existing.get('user_id'),
                    "name": existing.get('name') or '',
                    "age": existing.get('age'),
                    "gender": existing.get('gender') or '',
                    "role": existing.get('role') or 'user',
                    "birthDate": existing.get('birth_date') or ''
                }
            })
        else:
            # 新用户：这个微信号还没有在 users 表里注册过，

            # 前端拿着这个 openid，走正常的注册（患者）或绑定（家属/医生）流程
            return jsonify({
                "code": 0,
                "msg": "success",
                "data": {
                    "isNewUser": True,
                    "openid": openid
                }
            })

    except Exception as e:
        print(f"❌ [wx_login] 查询用户失败: {e}", flush=True)
        return jsonify({"code": -1, "msg": f"查询用户信息失败: {str(e)}"}), 500


# ============================================================
# ★ 新增：取消绑定功能（三方均可发起）
# 患者可以移除某个家属/医生的查看权限；家属/医生可以主动退出/结束关系。
# 不做物理删除，而是把 status 改成 'cancelled'，保留历史记录方便审计。
# 身份校验用 openid 反查操作者真实 user_id，确认这个人确实是这段绑定关系
# 里的一方（patient_id 或 family_id/doctor_id 之一），才允许操作——
# 不能让任意一方凭空传两个陌生人的 ID 就把别人的绑定关系解除掉。
# 取消后另一方的提醒通知，计划放进未来的 feedback 模块里做，这里不涉及。
# ============================================================

@binding_bp.route('/get_my_bindings', methods=['GET'])
def get_my_bindings():
    """
    患者查看自己当前绑定的家属列表 + 医生列表（供 more 页面"管理绑定"功能使用）。
    用 openid 确认这是患者本人在查询，而不是任何人传个 userId 就能看别人的绑定关系。
    """
    openid = request.args.get('openid')
    if not openid:
        return jsonify({"code": 1, "msg": "缺少 openid"}), 400

    conn = database.get_db()
    cursor = conn.cursor()
    ph = database.get_placeholder()

    try:
        cursor.execute(f"SELECT user_id FROM users WHERE openid = {ph}", (openid,))
        me = cursor.fetchone()
        if not me:
            return jsonify({"code": -1, "msg": "找不到对应的用户，请先完成注册"}), 404
        my_user_id = dict(me)['user_id']

        cursor.execute(
            f"SELECT family_id, name, created_at FROM family_bindings WHERE patient_id = {ph} AND status = 'active'",
            (my_user_id,)
        )
        family_rows = cursor.fetchall()
        family_list = []
        for row in family_rows:
            row = dict(row)
            family_list.append({
                "viewerId": row['family_id'],
                "name": row.get('name') or '家属用户',
                "boundAt": str(row.get('created_at') or '')
            })

        cursor.execute(
            f"SELECT doctor_id, doctor_name, hospital, department, created_at FROM doctor_bindings WHERE patient_id = {ph} AND status = 'active'",
            (my_user_id,)
        )
        doctor_rows = cursor.fetchall()
        doctor_list = []
        for row in doctor_rows:
            row = dict(row)
            doctor_list.append({
                "viewerId": row['doctor_id'],
                "name": row.get('doctor_name') or '医生用户',
                "hospital": row.get('hospital') or '',
                "department": row.get('department') or '',
                "boundAt": str(row.get('created_at') or '')
            })

        return jsonify({
            "code": 0,
            "msg": "success",
            "data": {
                "familyList": family_list,
                "doctorList": doctor_list
            }
        })

    except Exception as e:
        print(f"❌ [get_my_bindings] 查询失败: {e}", flush=True)
        return jsonify({"code": -1, "msg": f"查询失败: {str(e)}"}), 500


@binding_bp.route('/cancel_binding', methods=['POST'])
def cancel_binding():
    """
    取消一段绑定关系。患者、家属、医生三方均可发起。
    请求体: { openid, bindingType: 'family'|'doctor', patientId, viewerId }
    校验：用 openid 反查操作者真实 user_id，必须等于 patientId 或 viewerId 之一，
    否则拒绝（防止陌生人取消不相干的两个人之间的绑定关系）。
    """
    data = request.json or {}
    openid = data.get('openid')
    binding_type = data.get('bindingType')
    patient_id = data.get('patientId')
    viewer_id = data.get('viewerId')

    if not openid or not binding_type or not patient_id or not viewer_id:
        return jsonify({"code": -1, "msg": "缺少必要参数"}), 400
    if binding_type not in ('family', 'doctor'):
        return jsonify({"code": -1, "msg": "bindingType 参数不合法"}), 400

    conn = database.get_db()
    cursor = conn.cursor()
    ph = database.get_placeholder()

    try:
        cursor.execute(f"SELECT user_id FROM users WHERE openid = {ph}", (openid,))
        me = cursor.fetchone()
        if not me:
            return jsonify({"code": -1, "msg": "找不到对应的用户"}), 404
        my_user_id = dict(me)['user_id']

        # ★ 权限校验：操作者必须是这段关系里的一方（患者本人，或对应的家属/医生本人）
        if my_user_id != patient_id and my_user_id != viewer_id:
            print(f"⚠️ [cancel_binding] 权限拒绝: {my_user_id} 试图取消 {patient_id}-{viewer_id} 的绑定", flush=True)
            return jsonify({"code": -1, "msg": "无权限操作这段绑定关系"}), 403

        table = 'family_bindings' if binding_type == 'family' else 'doctor_bindings'
        id_col = 'family_id' if binding_type == 'family' else 'doctor_id'

        cursor.execute(
            f"UPDATE {table} SET status = 'cancelled' WHERE {id_col} = {ph} AND patient_id = {ph}",
            (viewer_id, patient_id)
        )
        conn.commit()

        if cursor.rowcount == 0:
            return jsonify({"code": -1, "msg": "未找到这段绑定关系，可能已经取消过"}), 404

        print(f"🟢 [DB] 绑定已取消: {binding_type} {viewer_id} <-> patient {patient_id}, 操作者={my_user_id}", flush=True)
        return jsonify({"code": 0, "msg": "已取消绑定"})

    except Exception as e:
        print(f"❌ [cancel_binding] 取消失败: {e}", flush=True)
        return jsonify({"code": -1, "msg": f"取消失败: {str(e)}"}), 500


@binding_bp.route('/get_patient_doctors', methods=['GET'])
def get_patient_doctors():
    """
    获取某个患者当前绑定的全部医生列表——留言功能选"跟哪个医生的那条线"要用到。
    ★ 跟 get_my_bindings 的区别：get_my_bindings 是"用 openid 查我自己(患者)的绑定"，
    只有患者本人查自己能用；这个接口是"给定 patientId，查这个患者绑定了哪些医生"，
    患者本人或者这个患者的绑定家属都能查（家属要看患者绑定了哪些医生，才能选一条线发言）。
    """
    patient_id = request.args.get('patientId')
    viewer_id = request.args.get('viewerId')

    if not patient_id or not viewer_id:
        return jsonify({"code": 1, "msg": "缺少必要参数(patientId/viewerId)"}), 400

    conn = database.get_db()
    cursor = conn.cursor()
    ph = database.get_placeholder()

    try:
        # 权限：患者查自己，或者对这个患者有 active 家属绑定，才能看
        if viewer_id != patient_id:
            cursor.execute(
                f"SELECT id FROM family_bindings WHERE family_id={ph} AND patient_id={ph} AND status='active'",
                (viewer_id, patient_id)
            )
            if not cursor.fetchone():
                return jsonify({"code": 1, "msg": "无权限查看该患者的绑定医生列表"}), 403

        cursor.execute(
            f"SELECT doctor_id, doctor_name, hospital, department FROM doctor_bindings "
            f"WHERE patient_id={ph} AND status='active'",
            (patient_id,)
        )
        rows = [dict(r) for r in cursor.fetchall()]
        doctors = [{
            "doctorId": r['doctor_id'],
            "doctorName": r.get('doctor_name') or '医生',
            "hospital": r.get('hospital') or '',
            "department": r.get('department') or ''
        } for r in rows]

        return jsonify({"code": 0, "data": doctors})
    except Exception as e:
        print(f"❌ [get_patient_doctors] 查询失败: {e}", flush=True)
        return jsonify({"code": 1, "msg": f"查询失败: {str(e)}"}), 500