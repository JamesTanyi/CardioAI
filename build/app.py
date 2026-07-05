#!/usr/bin/env python
"""
BloodTrack CloudRun 主入口
"""

import os
import sys
import json
from flask import Flask, request, jsonify
from datetime import datetime

# 数据库配置
# 优先级: 强制 SQLite > 环境变量 USE_CLOUD_DB > 默认 MySQL
_force_sqlite = os.environ.get("FORCE_SQLITE", "").lower() == "true"
_use_cloud_db_env = os.environ.get("USE_CLOUD_DB", "true").lower() == "true"
USE_CLOUD_DB = _use_cloud_db_env and not _force_sqlite

if USE_CLOUD_DB:
    import pymysql
    DB_CONFIG = {
        'host': os.environ.get("DB_HOST", "10.0.0.100"),
        'port': int(os.environ.get("DB_PORT", 3306)),
        'user': os.environ.get("DB_USER", "root"),
        'password': os.environ.get("DB_PASSWORD", ""),
        'database': os.environ.get("DB_NAME", "cardioai"),
        'charset': 'utf8mb4',
        'connect_timeout': 5
    }
    print("🔄 尝试连接腾讯云 MySQL 数据库...", flush=True)
    
    def _get_cloud_db():
        conn = pymysql.connect(**DB_CONFIG)
        conn.cursorclass = pymysql.cursors.DictCursor
        return conn
    
    def _init_cloud_db():
        conn = _get_cloud_db()
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS measurements (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                user_id     VARCHAR(100) NOT NULL,
                sbp         INT NOT NULL,
                dbp         INT NOT NULL,
                hr          INT DEFAULT 75,
                symptoms    TEXT DEFAULT '[]',
                risk_level  VARCHAR(20) DEFAULT 'normal',
                risk_text   TEXT DEFAULT '',
                analysis    TEXT DEFAULT '{}',
                datetime    VARCHAR(50) NOT NULL,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_user_id (user_id),
                INDEX idx_datetime (datetime)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                user_id     VARCHAR(100) UNIQUE NOT NULL,
                name        VARCHAR(50) DEFAULT '',
                age         INT DEFAULT 0,
                gender      VARCHAR(10) DEFAULT '',
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS family_bindings (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                family_id   VARCHAR(100) NOT NULL,
                patient_id  VARCHAR(100) NOT NULL,
                name        VARCHAR(50) NOT NULL,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY unique_binding (family_id, patient_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS feedbacks (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                from_id     VARCHAR(100) NOT NULL,
                from_role   VARCHAR(20) NOT NULL,
                to_id       VARCHAR(100) NOT NULL,
                content     TEXT NOT NULL,
                is_read     INT DEFAULT 0,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_to_id (to_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS doctor_bindings (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                doctor_id   VARCHAR(100) NOT NULL,
                patient_id  VARCHAR(100) NOT NULL,
                doctor_name VARCHAR(50) DEFAULT '',
                hospital    VARCHAR(200) DEFAULT '',
                department  VARCHAR(100) DEFAULT '',
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY unique_dr_binding (doctor_id, patient_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS invite_codes (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                code        VARCHAR(10) NOT NULL UNIQUE,
                patient_id  VARCHAR(100) NOT NULL,
                used        INT DEFAULT 0,
                used_by     VARCHAR(100) DEFAULT '',
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at  TIMESTAMP NULL,
                INDEX idx_code (code),
                INDEX idx_patient (patient_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
        conn.close()
        print("✅ [DB] MySQL 初始化完成", flush=True)
    
    # 尝试连接 MySQL，失败则自动降级到 SQLite
    try:
        _init_cloud_db()
        get_db = _get_cloud_db
        init_db = _init_cloud_db
    except Exception as mysql_err:
        print(f"❌ MySQL 连接失败: {mysql_err}", flush=True)
        print("🔽 自动降级到 SQLite 模式", flush=True)
        USE_CLOUD_DB = False

if not USE_CLOUD_DB:
    import sqlite3
    DB_PATH = os.environ.get("DB_PATH", "/tmp/bloodtrack.db")
    print(f"⚠️ 使用本地 SQLite 数据库: {DB_PATH}", flush=True)
    
    def get_db():
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        return conn
    
    def init_db():
        conn = get_db()
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS measurements (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id     TEXT NOT NULL,
                sbp         INTEGER NOT NULL,
                dbp         INTEGER NOT NULL,
                hr          INTEGER DEFAULT 75,
                symptoms    TEXT DEFAULT '[]',
                risk_level  TEXT DEFAULT 'normal',
                risk_text   TEXT DEFAULT '',
                analysis    TEXT DEFAULT '{}',
                datetime    TEXT NOT NULL,
                created_at  TEXT DEFAULT (datetime('now'))
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id     TEXT UNIQUE NOT NULL,
                name        TEXT DEFAULT '',
                age         INTEGER DEFAULT 0,
                gender      TEXT DEFAULT '',
                created_at  TEXT DEFAULT (datetime('now'))
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS family_bindings (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                family_id   TEXT NOT NULL,
                patient_id  TEXT NOT NULL,
                name        TEXT NOT NULL,
                created_at  TEXT DEFAULT (datetime('now')),
                UNIQUE(family_id, patient_id)
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS feedbacks (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                from_id     TEXT NOT NULL,
                from_role   TEXT NOT NULL,
                to_id       TEXT NOT NULL,
                content     TEXT NOT NULL,
                is_read     INTEGER DEFAULT 0,
                created_at  TEXT DEFAULT (datetime('now'))
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS doctor_bindings (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                doctor_id   TEXT NOT NULL,
                patient_id  TEXT NOT NULL,
                doctor_name TEXT DEFAULT '',
                hospital    TEXT DEFAULT '',
                department  TEXT DEFAULT '',
                created_at  TEXT DEFAULT (datetime('now')),
                UNIQUE(doctor_id, patient_id)
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS invite_codes (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                code        TEXT NOT NULL UNIQUE,
                patient_id  TEXT NOT NULL,
                used        INTEGER DEFAULT 0,
                used_by     TEXT DEFAULT '',
                created_at  TEXT DEFAULT (datetime('now')),
                expires_at  TEXT
            )
        """)
        conn.commit()
        conn.close()
        print("✅ [DB] SQLite 初始化完成", flush=True)

try:
    from engine.cardiovascular_engine import CardiovascularEngine
    EngineClass = CardiovascularEngine
    EngineError = None
except Exception as e:
    print(f"❌ 警告: 无法导入 CardiovascularEngine: {e}", flush=True)
    EngineClass = None
    EngineError = str(e)

def _normalize_record_time(rec):
    if rec is None:
        return rec
    rec = dict(rec)
    ts = rec.get("datetime") or rec.get("timestamp") or rec.get("date")
    if isinstance(ts, str):
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M"):
            try:
                rec["datetime"] = datetime.strptime(ts, fmt)
                break
            except ValueError:
                continue
    if "datetime" not in rec or not isinstance(rec["datetime"], datetime):
        rec["datetime"] = datetime.now()
    if "sbp" in rec and "dbp" in rec:
        rec["pp"] = rec.get("pp", rec["sbp"] - rec["dbp"])
    elif "pp" not in rec:
        rec["pp"] = 40
    if "hr" not in rec:
        rec["hr"] = 70
    return rec


def _format_record_for_db(item):
    return {
        "user_id": item.get("userId") or item.get("user_id"),
        "sbp": int(item.get("sbp", 0)),
        "dbp": int(item.get("dbp", 0)),
        "hr": int(item.get("hr", 75)),
        "symptoms": item.get("symptoms", []),
        "risk_level": item.get("riskLevel") or item.get("risk_level") or "normal",
        "risk_text": item.get("riskText") or item.get("risk_text") or "",
        "analysis": item.get("analysis") or {},
        "datetime": item.get("datetime") or item.get("date") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }


def _save_measurement(conn, cursor, item):
    record = _format_record_for_db(item)
    if not all([record["user_id"], record["sbp"], record["dbp"], record["datetime"]]):
        raise ValueError("Missing required record fields")
    if USE_CLOUD_DB:
        cursor.execute("INSERT IGNORE INTO users (user_id) VALUES (%s)", (record["user_id"],))
        cursor.execute(
            """
                INSERT INTO measurements
                    (user_id, sbp, dbp, hr, symptoms, risk_level, risk_text, analysis, datetime)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                record["user_id"],
                record["sbp"], record["dbp"], record["hr"],
                json.dumps(record["symptoms"], ensure_ascii=False),
                record["risk_level"], record["risk_text"],
                json.dumps(record["analysis"], ensure_ascii=False),
                record["datetime"]
            )
        )
    else:
        cursor.execute("INSERT OR IGNORE INTO users (user_id) VALUES (?)", (record["user_id"],))
        cursor.execute(
            """
                INSERT INTO measurements
                    (user_id, sbp, dbp, hr, symptoms, risk_level, risk_text, analysis, datetime)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                record["user_id"],
                record["sbp"], record["dbp"], record["hr"],
                json.dumps(record["symptoms"], ensure_ascii=False),
                record["risk_level"], record["risk_text"],
                json.dumps(record["analysis"], ensure_ascii=False),
                record["datetime"]
            )
        )


def _fetch_history_from_db(user_id, limit=90):
    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT * FROM measurements WHERE user_id=%s ORDER BY datetime DESC LIMIT %s",
                (user_id, limit)
            )
        else:
            cursor.execute(
                "SELECT * FROM measurements WHERE user_id = ? ORDER BY datetime DESC LIMIT ?",
                (user_id, limit)
            )
        rows = cursor.fetchall()
        results = []
        for row in rows:
            rec = dict(row) if isinstance(row, dict) else dict(row)
            rec["symptoms"] = json.loads(rec.get("symptoms") or "[]")
            rec["analysis"] = json.loads(rec.get("analysis") or "{}")
            results.append({
                "userId": rec.get("user_id"),
                "sbp": rec.get("sbp"),
                "dbp": rec.get("dbp"),
                "hr": rec.get("hr"),
                "symptoms": rec.get("symptoms"),
                "riskLevel": rec.get("risk_level"),
                "riskText": rec.get("risk_text"),
                "analysis": rec.get("analysis"),
                "datetime": rec.get("datetime")
            })
        return results
    finally:
        conn.close()

app = Flask(__name__)
init_db()

@app.route("/", methods=["GET"])
def health():
    return "Python service is running"

# ──────────────────────────────────────────────
# /analyze  核心分析
# ──────────────────────────────────────────────
@app.route("/analyze", methods=["POST"])
def analyze():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    print("📥 [Request] 收到 /analyze 请求", flush=True)
    if not data:
        return jsonify({"error": "Empty request body"}), 400
    if EngineClass is None:
        return jsonify({"code": -1, "error": "Engine load failed", "detail": EngineError}), 500

    history = data.get("history", [])
    current = data.get("current")
    if current is None:
        return jsonify({"error": "Missing 'current' record"}), 400
    if history is None:
        history = []
    if not isinstance(history, list):
        return jsonify({"error": "'history' must be a list"}), 400

    if not history:
        fallback_user_id = current.get("userId") or current.get("user_id")
        if fallback_user_id:
            history = _fetch_history_from_db(fallback_user_id, limit=90)

    history = [_normalize_record_time(r) for r in history]
    current = _normalize_record_time(current)

    try:
        engine = EngineClass(history, current)
        result = engine.run_all_diagnostics()
        print(f"✅ [Engine] 风险等级: {result.get('risk_level')}", flush=True)
        return jsonify({"code": 0, "data": result})
    except Exception as e:
        return jsonify({"error": "Engine execution failed", "detail": str(e)}), 500

# ──────────────────────────────────────────────
# /save_history  保存测量记录
# ──────────────────────────────────────────────
@app.route("/save_history", methods=["POST"])
def save_history():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    records = []
    if isinstance(data, dict) and data.get("history") and isinstance(data.get("history"), list):
        records = data.get("history")
    elif isinstance(data, list):
        records = data
    else:
        records = [data]

    conn = get_db()
    cursor = conn.cursor()
    saved = 0
    try:
        for item in records:
            try:
                _save_measurement(conn, cursor, item)
                saved += 1
            except Exception as inner_e:
                print(f"⚠️ 跳过无效记录: {inner_e}", flush=True)

        conn.commit()
        return jsonify({"code": 0, "message": "保存成功", "saved": saved})
    except Exception as e:
        return jsonify({"error": "保存失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /get_history  读取历史（患者自己或已绑定家属/医生）
# ──────────────────────────────────────────────
@app.route("/get_history", methods=["GET"])
def get_history():
    user_id   = request.args.get("userId") or request.args.get("user_id")
    viewer_id = request.args.get("viewerId")
    limit     = int(request.args.get("limit", 90))

    if not user_id:
        return jsonify({"error": "缺少 userId 参数"}), 400

    if viewer_id and viewer_id != user_id:
        conn = get_db()
        cursor = conn.cursor()
        # 同时检查家属绑定和医生绑定
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT id FROM family_bindings WHERE family_id=%s AND patient_id=%s",
                (viewer_id, user_id)
            )
            family_binding = cursor.fetchone()
            if not family_binding:
                cursor.execute(
                    "SELECT id FROM doctor_bindings WHERE doctor_id=%s AND patient_id=%s",
                    (viewer_id, user_id)
                )
                doctor_binding = cursor.fetchone()
            else:
                doctor_binding = None
        else:
            cursor.execute(
                "SELECT id FROM family_bindings WHERE family_id=? AND patient_id=?",
                (viewer_id, user_id)
            )
            family_binding = cursor.fetchone()
            if not family_binding:
                cursor.execute(
                    "SELECT id FROM doctor_bindings WHERE doctor_id=? AND patient_id=?",
                    (viewer_id, user_id)
                )
                doctor_binding = cursor.fetchone()
            else:
                doctor_binding = None
        conn.close()
        if not family_binding and not doctor_binding:
            return jsonify({"error": "无权限查看该用户数据，请先绑定"}), 403

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute("""
                SELECT * FROM measurements
                WHERE user_id = %s
                ORDER BY datetime DESC
                LIMIT %s
            """, (user_id, limit))
        else:
            cursor.execute("""
                SELECT * FROM measurements
                WHERE user_id = ?
                ORDER BY datetime DESC
                LIMIT ?
            """, (user_id, limit))
        
        rows = cursor.fetchall()
        
        records = []
        for row in rows:
            rec = dict(row) if isinstance(row, dict) else dict(row)
            # MySQL 返回的是字符串，需要解析
            rec["symptoms"] = json.loads(rec.get("symptoms") or "[]")
            rec["analysis"] = json.loads(rec.get("analysis") or "{}")
            records.append(rec)

        return jsonify({"code": 0, "data": records})
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /bind_family  家属绑定患者
# ──────────────────────────────────────────────
@app.route("/bind_family", methods=["POST"])
def bind_family():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    family_id  = data.get("familyId")
    patient_id = data.get("patientId")
    name       = data.get("name", "家人")

    if not all([family_id, patient_id]):
        return jsonify({"error": "缺少 familyId 或 patientId"}), 400
    if family_id == patient_id:
        return jsonify({"error": "不能绑定自己"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute("SELECT id FROM users WHERE user_id=%s", (patient_id,))
        else:
            cursor.execute("SELECT id FROM users WHERE user_id=?", (patient_id,))
        user = cursor.fetchone()
        if not user:
            # 患者尚未向服务器同步过数据，自动注册
            if USE_CLOUD_DB:
                cursor.execute("INSERT IGNORE INTO users (user_id) VALUES (%s)", (patient_id,))
            else:
                cursor.execute("INSERT OR IGNORE INTO users (user_id) VALUES (?)", (patient_id,))
            print(f"📝 [DB] 自动创建患者账户: {patient_id}", flush=True)

        if USE_CLOUD_DB:
            cursor.execute("""
                INSERT INTO family_bindings (family_id, patient_id, name)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE name=VALUES(name)
            """, (family_id, patient_id, name))
        else:
            cursor.execute("""
                INSERT OR REPLACE INTO family_bindings (family_id, patient_id, name)
                VALUES (?, ?, ?)
            """, (family_id, patient_id, name))
        conn.commit()
        print(f"🔗 [DB] 绑定: {family_id} → {patient_id} ({name})", flush=True)
        return jsonify({"code": 0, "message": "绑定成功"})
    except Exception as e:
        return jsonify({"error": "绑定失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /get_family_list  获取已绑定家人列表
# ──────────────────────────────────────────────
@app.route("/get_family_list", methods=["GET"])
def get_family_list():
    family_id = request.args.get("familyId")
    if not family_id:
        return jsonify({"error": "缺少 familyId 参数"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute("""
                SELECT patient_id, name, created_at
                FROM family_bindings
                WHERE family_id = %s
                ORDER BY created_at DESC
            """, (family_id,))
        else:
            cursor.execute("""
                SELECT patient_id, name, created_at
                FROM family_bindings
                WHERE family_id = ?
                ORDER BY created_at DESC
            """, (family_id,))
        rows = cursor.fetchall()
        return jsonify({"code": 0, "data": [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /send_feedback  家属或医生发反馈给患者
# ──────────────────────────────────────────────
@app.route("/send_feedback", methods=["POST"])
def send_feedback():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    from_id   = data.get("fromId")
    from_role = data.get("fromRole", "family")
    to_id     = data.get("toId")
    content   = data.get("content", "").strip()

    if not all([from_id, to_id, content]):
        return jsonify({"error": "缺少 fromId / toId / content"}), 400
    if len(content) > 500:
        return jsonify({"error": "反馈内容不能超过500字"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        # 根据角色检查对应的绑定关系
        if from_role == "doctor":
            if USE_CLOUD_DB:
                cursor.execute(
                    "SELECT id FROM doctor_bindings WHERE doctor_id=%s AND patient_id=%s",
                    (from_id, to_id)
                )
            else:
                cursor.execute(
                    "SELECT id FROM doctor_bindings WHERE doctor_id=? AND patient_id=?",
                    (from_id, to_id)
                )
        else:
            if USE_CLOUD_DB:
                cursor.execute(
                    "SELECT id FROM family_bindings WHERE family_id=%s AND patient_id=%s",
                    (from_id, to_id)
                )
            else:
                cursor.execute(
                    "SELECT id FROM family_bindings WHERE family_id=? AND patient_id=?",
                    (from_id, to_id)
                )
        binding = cursor.fetchone()
        if not binding:
            return jsonify({"error": "未绑定该患者，无法发送反馈"}), 403

        if USE_CLOUD_DB:
            cursor.execute("""
                INSERT INTO feedbacks (from_id, from_role, to_id, content)
                VALUES (%s, %s, %s, %s)
            """, (from_id, from_role, to_id, content))
        else:
            cursor.execute("""
                INSERT INTO feedbacks (from_id, from_role, to_id, content)
                VALUES (?, ?, ?, ?)
            """, (from_id, from_role, to_id, content))
        conn.commit()
        print(f"💬 [DB] 反馈: {from_id}({from_role}) → {to_id}", flush=True)
        return jsonify({"code": 0, "message": "反馈已发送"})
    except Exception as e:
        return jsonify({"error": "发送失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /get_feedback  患者查看收到的反馈
# ──────────────────────────────────────────────
@app.route("/get_feedback", methods=["GET"])
def get_feedback():
    user_id = request.args.get("userId")
    if not user_id:
        return jsonify({"error": "缺少 userId 参数"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute("""
                SELECT id, from_id, from_role, content, is_read, created_at
                FROM feedbacks
                WHERE to_id = %s
                ORDER BY created_at DESC
                LIMIT 50
            """, (user_id,))
        else:
            cursor.execute("""
                SELECT id, from_id, from_role, content, is_read, created_at
                FROM feedbacks
                WHERE to_id = ?
                ORDER BY created_at DESC
                LIMIT 50
            """, (user_id,))
        rows = cursor.fetchall()

        feedbacks = [dict(r) for r in rows]
        unread_count = sum(1 for f in feedbacks if f["is_read"] == 0)

        if USE_CLOUD_DB:
            cursor.execute(
                "UPDATE feedbacks SET is_read=1 WHERE to_id=%s AND is_read=0",
                (user_id,)
            )
        else:
            cursor.execute(
                "UPDATE feedbacks SET is_read=1 WHERE to_id=? AND is_read=0",
                (user_id,)
            )
        conn.commit()
        return jsonify({"code": 0, "data": feedbacks, "unread": unread_count})
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /bind_doctor  医生绑定患者
# ──────────────────────────────────────────────
@app.route("/bind_doctor", methods=["POST"])
def bind_doctor():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    doctor_id   = data.get("doctorId")
    patient_id  = data.get("patientId")
    doctor_name = data.get("doctorName", "")
    hospital    = data.get("hospital", "")
    department  = data.get("department", "")

    if not all([doctor_id, patient_id]):
        return jsonify({"error": "缺少 doctorId / patientId"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT id FROM users WHERE user_id=%s", (patient_id,)
            )
        else:
            cursor.execute(
                "SELECT id FROM users WHERE user_id=?", (patient_id,)
            )
        user = cursor.fetchone()
        if not user:
            # 患者尚未向服务器同步过数据，自动注册
            if USE_CLOUD_DB:
                cursor.execute("INSERT IGNORE INTO users (user_id) VALUES (%s)", (patient_id,))
            else:
                cursor.execute("INSERT OR IGNORE INTO users (user_id) VALUES (?)", (patient_id,))
            print(f"📝 [DB] 自动创建患者账户: {patient_id}", flush=True)

        if USE_CLOUD_DB:
            cursor.execute("""
                INSERT INTO doctor_bindings (doctor_id, patient_id, doctor_name, hospital, department)
                VALUES (%s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE doctor_name=VALUES(doctor_name),
                                        hospital=VALUES(hospital),
                                        department=VALUES(department)
            """, (doctor_id, patient_id, doctor_name, hospital, department))
        else:
            cursor.execute("""
                INSERT OR REPLACE INTO doctor_bindings (doctor_id, patient_id, doctor_name, hospital, department)
                VALUES (?, ?, ?, ?, ?)
            """, (doctor_id, patient_id, doctor_name, hospital, department))
        conn.commit()
        print(f"🩺 [DB] 医生绑定: {doctor_id} → {patient_id} ({doctor_name})", flush=True)
        return jsonify({"code": 0, "message": "绑定成功"})
    except Exception as e:
        return jsonify({"error": "绑定失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /generate_invite_code  患者生成邀请码（用于医生/家人绑定）
# ──────────────────────────────────────────────
@app.route("/generate_invite_code", methods=["POST"])
def generate_invite_code():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    patient_id = data.get("patientId") or data.get("user_id")
    if not patient_id:
        return jsonify({"error": "缺少 patientId"}), 400

    import random
    import string

    conn = get_db()
    cursor = conn.cursor()
    try:
        # 生成 6 位随机数字
        code = ''.join(random.choices(string.digits, k=6))

        if USE_CLOUD_DB:
            cursor.execute("""
                INSERT INTO invite_codes (code, patient_id, expires_at)
                VALUES (%s, %s, DATE_ADD(NOW(), INTERVAL 24 HOUR))
            """, (code, patient_id))
        else:
            cursor.execute("""
                INSERT INTO invite_codes (code, patient_id, expires_at)
                VALUES (?, ?, datetime('now', '+24 hours'))
            """, (code, patient_id))
        conn.commit()
        print(f"📨 [DB] 生成邀请码: {code} → {patient_id}", flush=True)
        return jsonify({"code": 0, "data": {"inviteCode": code, "expiresIn": "24小时"}})
    except Exception as e:
        return jsonify({"error": "生成失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /bind_doctor_by_code  医生通过邀请码绑定患者
# ──────────────────────────────────────────────
@app.route("/bind_doctor_by_code", methods=["POST"])
def bind_doctor_by_code():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({"error": "Invalid JSON", "detail": str(e)}), 400

    invite_code = data.get("code", "").strip()
    doctor_id   = data.get("doctorId")
    doctor_name = data.get("doctorName", "")
    hospital    = data.get("hospital", "")
    department  = data.get("department", "")

    if not invite_code or not doctor_id:
        return jsonify({"error": "缺少 code / doctorId"}), 400
    if len(invite_code) != 6 or not invite_code.isdigit():
        return jsonify({"error": "邀请码格式不正确，应为6位数字"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT * FROM invite_codes WHERE code=%s",
                (invite_code,)
            )
        else:
            cursor.execute(
                "SELECT * FROM invite_codes WHERE code=?",
                (invite_code,)
            )
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "邀请码不存在"}), 404

        row = dict(row)
        if row.get("used"):
            return jsonify({"error": "邀请码已被使用"}), 400

        # 检查是否过期
        expires_str = row.get("expires_at")
        if expires_str:
            from datetime import datetime as dt
            expires_at = dt.strptime(str(expires_str), "%Y-%m-%d %H:%M:%S")
            if dt.now() > expires_at:
                return jsonify({"error": "邀请码已过期，请重新生成"}), 400

        patient_id = row["patient_id"]

        # 防止自己绑定自己
        if doctor_id == patient_id:
            return jsonify({"error": "不能绑定自己"}), 400

        # 写入医生绑定表
        if USE_CLOUD_DB:
            cursor.execute("""
                INSERT INTO doctor_bindings (doctor_id, patient_id, doctor_name, hospital, department)
                VALUES (%s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE doctor_name=VALUES(doctor_name),
                                        hospital=VALUES(hospital),
                                        department=VALUES(department)
            """, (doctor_id, patient_id, doctor_name, hospital, department))
            # 标记邀请码已使用
            cursor.execute(
                "UPDATE invite_codes SET used=1, used_by=%s WHERE code=%s",
                (doctor_id, invite_code)
            )
        else:
            cursor.execute("""
                INSERT OR REPLACE INTO doctor_bindings (doctor_id, patient_id, doctor_name, hospital, department)
                VALUES (?, ?, ?, ?, ?)
            """, (doctor_id, patient_id, doctor_name, hospital, department))
            cursor.execute(
                "UPDATE invite_codes SET used=1, used_by=? WHERE code=?",
                (doctor_id, invite_code)
            )
        conn.commit()
        print(f"🩺 [DB] 医生通过邀请码绑定: {doctor_id} → {patient_id} (code: {invite_code})", flush=True)
        return jsonify({"code": 0, "message": "绑定成功"})
    except Exception as e:
        return jsonify({"error": "绑定失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /get_doctor_patients  医生查看已绑定患者（支持分页+搜索）
# ──────────────────────────────────────────────
@app.route("/get_doctor_patients", methods=["GET"])
def get_doctor_patients():
    doctor_id = request.args.get("doctorId")
    if not doctor_id:
        return jsonify({"error": "缺少 doctorId 参数"}), 400

    # 分页参数
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("pageSize", 20))
    keyword = request.args.get("keyword", "").strip()
    offset = (page - 1) * page_size

    conn = get_db()
    cursor = conn.cursor()
    try:
        # 构建查询条件
        where_clause = "WHERE doctor_id = %s" if USE_CLOUD_DB else "WHERE doctor_id = ?"
        params = [doctor_id] if USE_CLOUD_DB else [doctor_id]

        # 关键词搜索（患者ID模糊匹配）
        if keyword:
            where_clause += " AND patient_id LIKE %s" if USE_CLOUD_DB else " AND patient_id LIKE ?"
            keyword_pattern = f"%{keyword}%"
            params.append(keyword_pattern)

        # 查询总数
        count_sql = f"SELECT COUNT(*) as total FROM doctor_bindings {where_clause}"
        cursor.execute(count_sql, tuple(params) if USE_CLOUD_DB else params)
        total = cursor.fetchone()["total"]

        # 查询分页数据
        order_clause = "ORDER BY created_at DESC"
        limit_clause = "LIMIT %s OFFSET %s" if USE_CLOUD_DB else "LIMIT ? OFFSET ?"
        params_with_limit = params + [page_size, offset]

        data_sql = f"""
            SELECT patient_id, doctor_name, hospital, department, created_at
            FROM doctor_bindings
            {where_clause}
            {order_clause}
            {limit_clause}
        """
        cursor.execute(data_sql, tuple(params_with_limit) if USE_CLOUD_DB else params_with_limit)
        rows = cursor.fetchall()

        return jsonify({
            "code": 0,
            "data": [dict(r) for r in rows],
            "total": total,
            "page": page,
            "pageSize": page_size,
            "hasMore": offset + page_size < total
        })
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /get_family_patient  家属查找绑定的患者
# ──────────────────────────────────────────────
@app.route("/get_family_patient", methods=["GET"])
def get_family_patient():
    viewer_id = request.args.get("viewerId", "").strip()
    if not viewer_id:
        return jsonify({"error": "缺少 viewerId 参数"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        if USE_CLOUD_DB:
            cursor.execute("""
                SELECT patient_id, name, created_at
                FROM family_bindings
                WHERE family_id = %s
                ORDER BY created_at DESC
                LIMIT 1
            """, (viewer_id,))
        else:
            cursor.execute("""
                SELECT patient_id, name, created_at
                FROM family_bindings
                WHERE family_id = ?
                ORDER BY created_at DESC
                LIMIT 1
            """, (viewer_id,))
        row = cursor.fetchone()
        if row:
            name_val = row["name"] if "name" in row.keys() else (row["patient_id"] or "")
            return jsonify({
                "code": 0,
                "patientId": row["patient_id"],
                "patientName": name_val or row["patient_id"],
                "relation": name_val or "家人"
            })
        else:
            return jsonify({"code": 0, "patientId": "", "patientName": "", "message": "未找到绑定"})
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /get_binding_status  检查用户绑定状态 + 预警摘要（APP式启动接口）
# 前端每次启动/恢复时调用此接口，作为所有绑定数据的唯一真相源
# ──────────────────────────────────────────────
@app.route("/get_binding_status", methods=["GET"])
def get_binding_status():
    user_id = request.args.get("userId", "").strip()
    if not user_id:
        return jsonify({"error": "缺少 userId 参数"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        result = {
            "hasFamilyBinding": False,
            "hasDoctorBinding": False,
            "familyPatients": [],           # 所有家属绑定的患者
            "doctorPatients": [],           # 所有医生绑定的患者（含风险）
            "familyAlertRisk": "none",      # 家属侧最高风险
            "doctorAlertCount": 0,          # 医生侧需关注人数
            "familyAlertSummary": None,     # 家属预警摘要
            "doctorAlertSummary": None      # 医生预警摘要
        }

        # ── 1. 检查家属绑定（支持多患者） ──
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT patient_id, name, created_at FROM family_bindings WHERE family_id=%s ORDER BY created_at DESC",
                (user_id,)
            )
        else:
            cursor.execute(
                "SELECT patient_id, name, created_at FROM family_bindings WHERE family_id=? ORDER BY created_at DESC",
                (user_id,)
            )
        family_rows = cursor.fetchall()
        if family_rows:
            result["hasFamilyBinding"] = True
            family_patients = []
            max_risk = "none"
            for row in family_rows:
                r = dict(row) if not isinstance(row, dict) else row
                pid = r.get("patient_id") or r["patient_id"]
                pname = r.get("name", "") or pid
                # 查该患者最新测量
                if USE_CLOUD_DB:
                    cursor.execute(
                        "SELECT sbp, dbp, hr, risk_level, datetime FROM measurements WHERE user_id=%s ORDER BY datetime DESC LIMIT 1",
                        (pid,)
                    )
                else:
                    cursor.execute(
                        "SELECT sbp, dbp, hr, risk_level, datetime FROM measurements WHERE user_id=? ORDER BY datetime DESC LIMIT 1",
                        (pid,)
                    )
                meas = cursor.fetchone()
                risk = "none"; sbp = None; dbp = None; hr = None; d = None
                if meas:
                    sbp = int(meas["sbp"] or 0)
                    dbp = int(meas["dbp"] or 0)
                    hr = int(meas["hr"] or 0)
                    rl = meas["risk_level"] if "risk_level" in meas.keys() else ""
                    d = str(meas["datetime"])[:10]
                    if rl == "high" or sbp >= 160 or dbp >= 100:
                        risk = "high"
                    elif rl == "moderate" or sbp >= 140 or dbp >= 90:
                        risk = "moderate"
                    elif sbp > 0:
                        risk = "normal"
                if risk == "high" or (risk == "moderate" and max_risk != "high"):
                    max_risk = risk
                family_patients.append({
                    "patientId": pid,
                    "patientName": pname,
                    "risk": risk,
                    "sbp": sbp,
                    "dbp": dbp,
                    "hr": hr,
                    "date": d
                })

            result["familyPatients"] = family_patients
            result["familyAlertRisk"] = max_risk
            # 取第一个（最新绑定）做快速摘要
            fp = family_patients[0]
            result["familyAlertSummary"] = {
                "patientId": fp["patientId"],
                "patientName": fp["patientName"],
                "risk": fp["risk"],
                "sbp": fp["sbp"],
                "dbp": fp["dbp"],
                "date": fp["date"]
            }

        # ── 2. 检查医生绑定（全量患者+风险） ──
        if USE_CLOUD_DB:
            cursor.execute(
                "SELECT patient_id, doctor_name, hospital, department, created_at FROM doctor_bindings WHERE doctor_id=%s ORDER BY created_at DESC",
                (user_id,)
            )
        else:
            cursor.execute(
                "SELECT patient_id, doctor_name, hospital, department, created_at FROM doctor_bindings WHERE doctor_id=? ORDER BY created_at DESC",
                (user_id,)
            )
        doctor_rows = cursor.fetchall()
        if doctor_rows:
            result["hasDoctorBinding"] = True
            doctor_patients = []
            high_count = moderate_count = normal_count = unmonitored_count = 0
            for row in doctor_rows:
                r = dict(row) if not isinstance(row, dict) else row
                pid = r.get("patient_id") or r["patient_id"]
                dname = r.get("doctor_name", "") or ""
                hosp = r.get("hospital", "") or ""
                dept = r.get("department", "") or ""
                if USE_CLOUD_DB:
                    cursor.execute(
                        "SELECT sbp, dbp, hr, risk_level, datetime FROM measurements WHERE user_id=%s ORDER BY datetime DESC LIMIT 1",
                        (pid,)
                    )
                else:
                    cursor.execute(
                        "SELECT sbp, dbp, hr, risk_level, datetime FROM measurements WHERE user_id=? ORDER BY datetime DESC LIMIT 1",
                        (pid,)
                    )
                m = cursor.fetchone()
                risk = "none"; sbp = None; dbp = None; hr = None; d = None
                if not m:
                    unmonitored_count += 1
                    risk = "none"
                else:
                    sbp = int(m["sbp"] or 0)
                    dbp = int(m["dbp"] or 0)
                    hr = int(m["hr"] or 0)
                    rl = m["risk_level"] if "risk_level" in m.keys() else ""
                    d = str(m["datetime"])[:10]
                    if rl == "high" or sbp >= 160 or dbp >= 100:
                        high_count += 1; risk = "high"
                    elif rl == "moderate" or sbp >= 140 or dbp >= 90:
                        moderate_count += 1; risk = "moderate"
                    elif sbp > 0:
                        normal_count += 1; risk = "normal"
                doctor_patients.append({
                    "patientId": pid,
                    "doctorName": dname,
                    "hospital": hosp,
                    "department": dept,
                    "risk": risk,
                    "sbp": sbp,
                    "dbp": dbp,
                    "hr": hr,
                    "date": d
                })

            result["doctorPatients"] = doctor_patients
            result["doctorAlertCount"] = high_count + moderate_count
            result["doctorAlertSummary"] = {
                "totalPatients": len(doctor_rows),
                "highRiskCount": high_count,
                "moderateRiskCount": moderate_count,
                "normalCount": normal_count,
                "unmonitoredCount": unmonitored_count
            }

        return jsonify({"code": 0, "data": result})
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /get_patients_risk_summary  医生查看患者风险汇总
# ──────────────────────────────────────────────
@app.route("/get_patients_risk_summary", methods=["GET"])
def get_patients_risk_summary():
    doctor_id = request.args.get("doctorId", "").strip()
    if not doctor_id:
        return jsonify({"error": "缺少 doctorId 参数"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        # 获取所有绑定患者
        if USE_CLOUD_DB:
            cursor.execute("""
                SELECT patient_id FROM doctor_bindings
                WHERE doctor_id = %s
            """, (doctor_id,))
        else:
            cursor.execute("""
                SELECT patient_id FROM doctor_bindings
                WHERE doctor_id = ?
            """, (doctor_id,))
        bindings = cursor.fetchall()
        patient_ids = [row["patient_id"] for row in bindings]

        if not patient_ids:
            return jsonify({"code": 0, "highRiskCount": 0, "moderateRiskCount": 0, "unmonitoredCount": 0, "details": {}})

        # 查询每个患者的最新测量
        risk_detail = {}
        high_count = moderate_count = unmonitored_count = 0

        for pid in patient_ids:
            if USE_CLOUD_DB:
                cursor.execute("""
                    SELECT sbp, dbp, datetime, risk_level FROM measurements
                    WHERE user_id = %s
                    ORDER BY datetime DESC
                    LIMIT 1
                """, (pid,))
            else:
                cursor.execute("""
                    SELECT sbp, dbp, datetime, risk_level FROM measurements
                    WHERE user_id = ?
                    ORDER BY datetime DESC
                    LIMIT 1
                """, (pid,))
            row = cursor.fetchone()
            if not row:
                unmonitored_count += 1
                risk_detail[pid] = {"risk": "none", "sbp": None, "dbp": None, "date": None}
                continue

            sbp = int(row["sbp"] or 0)
            dbp = int(row["dbp"] or 0)
            rl = row["risk_level"] if "risk_level" in row.keys() else ""

            if rl == "high" or sbp >= 160 or dbp >= 100:
                high_count += 1
                risk_detail[pid] = {"risk": "high", "sbp": sbp, "dbp": dbp, "date": str(row["datetime"])[:10]}
            elif rl == "moderate" or sbp >= 140 or dbp >= 90:
                moderate_count += 1
                risk_detail[pid] = {"risk": "moderate", "sbp": sbp, "dbp": dbp, "date": str(row["datetime"])[:10]}
            else:
                risk_detail[pid] = {"risk": "normal", "sbp": sbp, "dbp": dbp, "date": str(row["datetime"])[:10]}

        return jsonify({
            "code": 0,
            "highRiskCount": high_count,
            "moderateRiskCount": moderate_count,
            "unmonitoredCount": unmonitored_count,
            "details": risk_detail
        })
    except Exception as e:
        return jsonify({"error": "查询失败", "detail": str(e)}), 500
    finally:
        conn.close()

# ──────────────────────────────────────────────
# /upload_excel  上传并解析 Excel 文件
# ──────────────────────────────────────────────
@app.route("/upload_excel", methods=["POST"])
def upload_excel():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "没有文件"}), 400
        
        file = request.files['file']
        user_id = request.form.get('userId') or request.form.get('user_id')
        
        if not user_id:
            return jsonify({"error": "缺少 userId"}), 400
        
        if file.filename == '':
            return jsonify({"error": "文件名为空"}), 400
        
        print(f"📥 收到 Excel 文件: {file.filename}", flush=True)
        
        # 读取文件内容
        file_content = file.read()
        
        # 使用 openpyxl 库解析
        try:
            import io
            from openpyxl import load_workbook
            
            wb = load_workbook(io.BytesIO(file_content))
            sheet = wb.active
            
            # 获取表头（第一行）
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).strip() if cell.value else '')
            
            print(f"📋 Excel 表头: {headers}", flush=True)
            
            # 解析数据行
            records = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(v is None for v in row):
                    continue
                
                record = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        record[headers[col_idx]] = value
                
                records.append(record)
            
            print(f"📊 解析到 {len(records)} 条记录", flush=True)
            
        except ImportError:
            return jsonify({"error": "服务器未安装 xlsx 库"}), 500
        except Exception as e:
            return jsonify({"error": f"Excel 解析失败: {str(e)}"}), 500
        
        if not records:
            return jsonify({"code": 0, "message": "文件中没有数据", "saved": 0})
        
        # 保存到数据库
        conn = get_db()
        cursor = conn.cursor()
        saved = 0
        
        try:
            for item in records:
                try:
                    record = _format_record_for_db({**item, "userId": user_id})
                    if not all([record["user_id"], record["sbp"], record["dbp"], record["datetime"]]):
                        continue
                    
                    _save_measurement(conn, cursor, {**item, "userId": user_id})
                    saved += 1
                except Exception as inner_e:
                    print(f"⚠️ 跳过无效记录: {inner_e}", flush=True)
            
            conn.commit()
            print(f"✅ 成功保存 {saved} 条记录", flush=True)
            return jsonify({"code": 0, "message": "上传成功", "saved": saved})
        
        except Exception as e:
            return jsonify({"error": f"保存失败: {str(e)}"}), 500
        finally:
            conn.close()
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port, debug=False)
